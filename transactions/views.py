from email import message
from genericpath import samefile
from django.contrib import messages
from django.http.response import HttpResponseRedirect
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
import pandas as pd
import requests


from store.views import numOfDays
from .forms import *
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import *
from datetime import date
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

from django.urls import reverse
import csv
import mimetypes

import os

# Build paths inside the project like this: os.path.join(BASE_DIR, ...)
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


from datetime import datetime
import pytz

IST = pytz.timezone('Asia/Kolkata')




def demo(request):

    
    
    s = stock.objects.all()






from .filters import *
from django.db.models import Sum
from django.db.models import Count

@login_required(login_url='login')
def list_stock(request):

   
    # data = stock.objects.filter(quantity__gt=0).prefetch_related('product__project_material_re').order_by('product__category', 'product__grade')
    
    
    # total_stock = data.aggregate(total_stock=Sum('quantity'))['total_stock']


    data = product_qr.objects.filter(
        moved_to_scratch=False,
        moved_to_left_over=False,
        product__isnull=False
    ).values(
        'product', 
        'product__category__name', 
        'product__size__mm1', 
        'product__size__mm2', 
        'product__size__name', 
        'product__grade__name', 
        'product__thickness__name'
    ).annotate(total_quantity=Count('id')).order_by('product')

    # Prepare final data with related product_qr entries
    final_data = []
    
    for stock in data:

        # Get related product_qr entries for the specific product
        product_qr_entries = product_qr.objects.filter(product=stock['product'],  moved_to_scratch=False,
        moved_to_left_over=False,
        product__isnull=False)
        print('------------------------------------')
        print(product_qr_entries)

        # Append the stock data and related product_qr entries
        final_data.append({
            'stock': stock,
            'product_qr_entries': product_qr_entries
        })

    
    total_stock = data.aggregate(total_quantity1=Sum('total_quantity'))

    context = {
        'data': final_data,
        'total_stock' : total_stock,
        
    }

    return render(request, 'transactions/list_stock.html', context)


from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, ListFlowable, ListItem
)



from collections import defaultdict
from io import BytesIO
from datetime import datetime
import os

from django.core.mail import EmailMessage
from django.http import JsonResponse
from django.utils import timezone
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfgen import canvas

from .models import product_qr  # Replace with actual import
from django.conf import settings
from django.db.models import Count


def stock_report_email(request):
    today = timezone.now().date()

    data = product_qr.objects.filter(
        moved_to_scratch=False,
        moved_to_left_over=False,
        product__isnull=False
    ).values(
        'product',
        'product__category__name',
        'product__size__mm1',
        'product__size__mm2',
        'product__size__name',
        'product__grade__name',
        'product__thickness__name',
    ).annotate(total_quantity=Count('id')).order_by('product')

    product_qr_map = defaultdict(list)
    for qr in product_qr.objects.filter(
        moved_to_scratch=False,
        moved_to_left_over=False,
        product__isnull=False
    ).select_related('product'):
        product_qr_map[qr.product_id].append(qr)

    buffer = BytesIO()

    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawRightString(A4[0] - 20, A4[1] - 20, f"Date: {today.strftime('%d-%m-%Y')}")
        canvas.drawRightString(A4[0] - 20, 15, f"Page {doc.page}")
        canvas.restoreState()

    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=40,
        bottomMargin=40
    )

    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1
    title_style.fontSize = 14

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        alignment=1,
        fontSize=12
    )

    elements.append(Paragraph("Ravi Raj Anodisers", title_style))
    elements.append(Paragraph("Stock Report", subtitle_style))
    elements.append(Spacer(1, 0.4 * cm))

    header = ["#", "Material Name", "Length MM", "Width MM", "Size", "Grade", "Thickness", "Quantity"]
    header_table = Table([header], colWidths=[30, 90, 60, 60, 80, 60, 60, 50])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2E86C1")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#AAB7B8")),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(header_table)

    total_quantity = 0

    for idx, stock in enumerate(data, 1):
        size = stock['product__size__name'] or f"{stock['product__size__mm1']} x {stock['product__size__mm2']}"
        table_row = [
            idx,
            stock['product__category__name'],
            stock['product__size__mm1'],
            stock['product__size__mm2'],
            size,
            stock['product__grade__name'],
            stock['product__thickness__name'],
            stock['total_quantity']
        ]
        total_quantity += stock['total_quantity']

        data_table = Table([table_row], colWidths=[30, 90, 60, 60, 80, 60, 60, 50])
        data_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#AAB7B8")),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
        ]))
        elements.append(data_table)

        qr_items = product_qr_map.get(stock['product'])
        if qr_items:
            qr_table_data = [["Finish", "Sheet No"]]
            for qr in qr_items:
                qr_table_data.append([str(qr.finish).upper(), str(qr.id)])

            qr_table = Table(qr_table_data, colWidths=[150, 100])
            qr_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#AAB7B8")),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#D6EAF8")),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
            ]))
            elements.append(qr_table)

        elements.append(Spacer(1, 0.2 * cm))

    total_row = ["", "", "", "", "", "", "Total Stock :-", total_quantity]
    total_table = Table([total_row], colWidths=[30, 90, 60, 60, 80, 60, 60, 50])
    total_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#AAB7B8")),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
    ]))
    elements.append(total_table)
    elements.append(Spacer(1, 0.5 * cm))

    footer_table = Table([
        ["Checked By:", "", "Signature:"],
        ["Verified By Name:", "", "Signature:"],
    ], colWidths=[120, 300, 100])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(footer_table)

    pdf.build(elements, onFirstPage=on_page, onLaterPages=on_page)

    # 🛠️ Write to file BEFORE closing buffer
    pdf_data = buffer.getvalue()
    file_path = os.path.join(settings.MEDIA_ROOT, 'stock_report.pdf')
    with open(file_path, 'wb') as f:
        f.write(pdf_data)
    buffer.close()

    # 📧 Email
    email = EmailMessage(
        subject='Stock Report PDF',
        body='Please find attached the Stock Report.',
        from_email='rradailyupdates@gmail.com',
        to=['pratikgosavi654@gmail.com'],  # Add others if needed
    )
    email.attach_file(file_path)
    email.send()

    return JsonResponse({'message': 'Stock report sent successfully'})



@login_required(login_url='login')
def list_left_over_stock(request):

   
    data = product_qr.objects.filter(
        moved_to_scratch=False,
        moved_to_left_over=True,
        product__isnull=False
    ).values(
        'product', 
        'product__category__name', 
        'product__size__name', 
        'product__grade__name', 
        'product__thickness__name'
    ).annotate(total_quantity=Count('id')).order_by('product')

    # Prepare final data with related product_qr entries
    final_data = []
    
    for stock in data:

        # Get related product_qr entries for the specific product
        product_qr_entries = product_qr.objects.filter(product=stock['product'], moved_to_scratch=False,
        moved_to_left_over=True,
        product__isnull=False)
        print('------------------------------------')
        print(product_qr_entries)

        # Append the stock data and related product_qr entries
        final_data.append({
            'stock': stock,
            'product_qr_entries': product_qr_entries
        })

    print(final_data)
    
    total_stock = data.aggregate(total_quantity1=Sum('total_quantity'))

    
    context = {
        'data': final_data,
        'total_stock' : total_stock,
        
    }

    return render(request, 'transactions/list_left_over_stock.html', context)

@login_required(login_url='login')
def list_dead_stock(request):

   
    data = product_qr.objects.filter(
        moved_to_scratch=True,
        moved_to_left_over=False,
        product__isnull=False
    ).values(
        'product', 
        'product__category__name', 
        'product__size__name', 
        'product__grade__name', 
        'product__thickness__name'
    ).annotate(total_quantity=Count('id')).order_by('product')

    # Prepare final data with related product_qr entries
    final_data = []
    
    for stock in data:

        # Get related product_qr entries for the specific product
        product_qr_entries = product_qr.objects.filter(product=stock['product'],  moved_to_scratch=True,
        moved_to_left_over=False,
        product__isnull=False)
        print('------------------------------------')
        print(product_qr_entries)

        # Append the stock data and related product_qr entries
        final_data.append({
            'stock': stock,
            'product_qr_entries': product_qr_entries
        })

    print(final_data)
    
    total_stock = data.aggregate(total_quantity1=Sum('total_quantity'))

    
    context = {
        'data': final_data,
        'total_stock' : total_stock,
        
    }


    return render(request, 'transactions/list_dead_stock.html', context)

@login_required(login_url='login')
def list_notifications(request):

   
    data = notification_table.objects.all()
    
    context = {
        'data': data,
        
    }

    return render(request, 'transactions/list_notifications.html', context)


@login_required(login_url='login')
def low_stock_report(request, notification_id):

   
    data = notification_table.objects.get(id = notification_id)
    
    data.is_reported = True
    data.save()

    return redirect('list_notifications')





@login_required(login_url='login')
def delete_images(request):

   
    data = product_qr.objects.filter(qr_code__isnull=False)

    for i in data:

        i.qr_code.delete()
   

    return redirect('list_generated_product_qr')



from datetime import datetime
import requests


def send_notification():


    # Define the URL
    url = "https://fcm.googleapis.com/fcm/send"

    # Define the payload
    payload = {
        "notification": {
            "body": "This is an FCM notification message!",
            "time": "FCM Message"
        },
        "registration_ids": [
            "deXBxhh5kaaduGWQn-XIWH:APA91bF6HOf_sdv0yLfgxKKhxr0V-ilVOzPtFEbU9MLhFyKISQDdOvFCVuA0dihoXvuB94iAXsBOOYWEz7i6o9u90e9NWDd3a5KeH8ANQQrnoFiXopO95-VBvM2X0b53EEIA3g4ZXV3l"
        ]
    }

    # Define the headers
    headers = {
        "Authorization": "key=AAAA-aJPhnk:APA91bGsRiq0bVsPc2DUPDWxiJvNOvdYpkcPG9pAT2V0aUzmNBws7O5uzdVOtCiY6d-7QrJ1PTGN2pTUenPIbL7ZB6qIwYomhsOHyaipHLlMT4dbFgKVTIb1yO6zsdkUC0aJLF7G7_8B",
        "Content-Type": "application/json"
    }

    # Send the POST request
    response = requests.post(url, json=payload, headers=headers)

    # Print the response
    print(response.status_code)
    print(response.text)





import http.client
import ssl 


access_token = "EAALeGznz5UwBO9cCf9mrwEd1vHBgB8neIziWXhS4AKGY02ZCVbfb5bTnSK7TCX6Qo1V0MZCHg7hNHQJYsNIZB17zlXaXFLv4HWJFWHZA0zeK57eZCClKyKxeAROKBh0kWB9PtjGbJeJsDWQSdqIjr20xrOBvk09nfWZCRn4xi5MTuyhco7C3U9P4OZBRbADDzLfKwZDZD"
# recipient_number = ["8237377298"]
recipient_number = ["9765054243", "8767515210", "8237377298"]
template_name = "qutation_added"
language_code = "en"





def send_low_stock_notification(request, token, recipient_number, template_name, language_code, dynamic_value):

    # for i in recipient_number:

    #     print(i)

    #     url = "https://graph.facebook.com/v20.0/363920080139942/messages"
    #     headers = {
    #         "Authorization": f"Bearer {token}",
    #         "Content-Type": "application/json"
    #     }
    #     payload = {
    #         "messaging_product": "whatsapp",
    #         "to": [i],
    #         "type": "template",
    #         "template": {
    #             "name": template_name,
    #             "language": {
    #                 "code": language_code
    #             },
    #             "components": [
    #                 {
    #                     "type": "body",
    #                     "parameters": [
    #                         {
    #                             "type": "text",
    #                             "text": dynamic_value
    #                         }
    #                     ]
    #                 }
    #             ]
    #         }
    #     }
        
    #     response = requests.post(url, headers=headers, data=json.dumps(payload))
    #     print(response)
    #     return response.json()



    email = EmailMessage(
            subject='Stock Alert',
            body=dynamic_value,
            from_email='rradailyupdates@gmail.com',
            to=['varad@ravirajanodisers.com', 'ravi@ravirajanodisers.com', 'raj@ravirajanodisers.com'],
            # to=['pratikgosavi654@gmail.com'],
        )

    email.send()

     


def work_alert(request, token, recipient_number, template_name, language_code, param1, param2):

    
    # url = "https://graph.facebook.com/v20.0/363920080139942/messages"
    # headers = {
    #     "Authorization": f"Bearer {token}",
    #     "Content-Type": "application/json"
    # }
    
    # for number in recipient_number:


    
    #     payload = {
    #         "messaging_product": "whatsapp",
    #         "to": number,
    #         "type": "template",
    #         "template": {
    #             "name": template_name,
    #             "language": {
    #                 "code": language_code
    #             },
    #             "components": [
    #                 {
    #                     "type": "body",
    #                     "parameters": [
    #                         {
    #                             "type": "text",
    #                             "text": param1
    #                         },
    #                         {
    #                             "type": "text",
    #                             "text": param2
    #                         }
    #                     ]
    #                 }
    #             ]
    #         }
    #     }
        
    #     response = requests.post(url, headers=headers, data=json.dumps(payload))

    #     print(response.json())

    email = EmailMessage(
            subject='Work Alert',
            body=param1,
            from_email='rradailyupdates@gmail.com',
            to=['varad@ravirajanodisers.com', 'ravi@ravirajanodisers.com', 'raj@ravirajanodisers.com', 'pratikgosavi654@gmail.com'],
            # to=['pratikgosavi654@gmail.com'],
        )

   
    # Send the email
    email.send()


        


        
def send_notification():


    # Define the URL
    url = "https://fcm.googleapis.com/fcm/send"

    # Define the payload
    payload = {
        "notification": {
            "body": "This is an FCM notification message!",
            "time": "FCM Message"
        },
        "registration_ids": [
            "deXBxhh5kaaduGWQn-XIWH:APA91bF6HOf_sdv0yLfgxKKhxr0V-ilVOzPtFEbU9MLhFyKISQDdOvFCVuA0dihoXvuB94iAXsBOOYWEz7i6o9u90e9NWDd3a5KeH8ANQQrnoFiXopO95-VBvM2X0b53EEIA3g4ZXV3l"
        ]
    }

    # Define the headers
    headers = {
        "Authorization": "key=AAAA-aJPhnk:APA91bGsRiq0bVsPc2DUPDWxiJvNOvdYpkcPG9pAT2V0aUzmNBws7O5uzdVOtCiY6d-7QrJ1PTGN2pTUenPIbL7ZB6qIwYomhsOHyaipHLlMT4dbFgKVTIb1yO6zsdkUC0aJLF7G7_8B",
        "Content-Type": "application/json"
    }

    # Send the POST request
    response = requests.post(url, json=payload, headers=headers)

    # Print the response
    print(response.status_code)
    print(response.text)





from django.forms.models import model_to_dict

import json
from django.forms.models import model_to_dict
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone

 
       

# views.py

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def subscribe(request):
    if request.method == 'POST':
        subscription_data = json.loads(request.body)
        # Store subscription_data in your database
        return JsonResponse({'success': True})





import csv



def downalo_data(request):

    projects_data = []
    projects_data.append(['Order ID', 'Customer', 'Sheet No', 'Previous Size', 'Used Size',	'Left Size', 'Updated At', 'Cutter'])

    for proj in project.objects.all():

        projects_data.append([proj.order_id, proj.customer, '', '', '',	'', '', ''])


        for material in project_material.objects.filter(project=proj):

            if material.sheet_no:

                try:
                    product_qr_instances = project_matarial_qr.objects.get(id = material.sheet_no)



                    a = material_history.objects.filter(product_qr = product_qr_instances.id, project = proj)


                    for z in a:
                        projects_data.append(['', '', z.product_qr.id, z.previous_size, z.used_size, z.left_size, z.updated_at, z.cutter])

                except project_matarial_qr.DoesNotExist:
                    
                    pass

    # Convert data to DataFrame
    df = pd.DataFrame(projects_data)

    # Export DataFrame to CSV
    df.to_csv('project_data.csv', index=False)




    


from django.http import HttpResponse
from django.conf import settings



from datetime import datetime, timedelta



def script(request):


    a = project.objects.filter(completed = False)

    message_body = ""

    for i in a:

        b = project_material.objects.filter(project = i)

        for z in b:

            product_qr_instance = product_qr.objects.get(id = z.sheet_no)

            current_date = datetime.now().date()
            two_days_ago = current_date - timedelta(days=2)

            # Check if the given date has exceeded 2 days from the current date
            if i.DC_date < two_days_ago:

                c = material_history.objects.filter(product_qr = product_qr_instance, project = i)

                if not c:

                    print('messae--------------------------------------')
                    
                    message_body += "Project Id: " + str(i.id) + " " + "Customer Name: " + str(i.customer.name) + " " + "Sheet no: " + str(z.sheet_no) + " Not updated from 2 days \n \n"

                            
    print(message_body)
    
    work_alert(request, access_token, recipient_number, 'sheet_update', language_code, message_body, 'z.sheet_no')



    return JsonResponse({'message': 'response message'}, status=400)






def sheet_tracking(request, sheet_id, rfid_value):


    sheets_rfid_instance = sheets_rifd.objects.get(sheet_id = sheet_id, status = True)

    sheet_tracking_history.objects.create(project_id = sheets_rfid_instance.project_id, sheet_id = sheets_rfid_instance.sheet_id)
    

    context = {

        'sheet_id' : sheet_id,
        'rfid_value' : rfid_value,

    }

    return JsonResponse(context)


def assign_rfid_to_sheet_reception_page(request, project_id):

    project_instance = project.objects.get(id = project_id)

    forms = project_Form(instance = project_instance)

    data_form = product_Form()

    data = project_material.objects.filter(project = project_instance)


    context = {
        'form': forms,
        'data_form': data_form,
        'data': data,
        'project_id': project_id,
    }
    return render(request, 'transactions/assign_rfid_to_sheet_reception_page.html', context)


from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.http import JsonResponse
from .models import sheets_rifd
import requests
import time

@csrf_exempt
def values_to_assign_rfid_to_sheet(request, project_id, sheet_id):
    
    node_endpoint = "http://192.168.66.98:80"

    try:
        # Make the request to the NodeMCU server
        response = requests.get(node_endpoint)

        # Print the content of the response
        print('Response content:', response.content)

        # Attempt to parse the response as JSON
        rfid_value = response.json()

        # Print the parsed JSON data
        print('Parsed JSON data:', rfid_value)

        # Serialize the response if needed
        serialized_response = json.dumps({'status': rfid_value})

        # Check if the response contains data
        if rfid_value:
            # Check if a sheet with the same project_id, sheet_id, and rfid_value exists
            check_for_active_sheets = sheets_rifd.objects.filter(project_id=project_id, sheet_id=sheet_id, rfid_value=rfid_value, status=True)
            if not check_for_active_sheets:
                sheets_rifd.objects.create(project_id=project_id, sheet_id=sheet_id, rfid_value=rfid_value)
                return JsonResponse(serialized_response, safe=False)
            else:
                return JsonResponse({'status': 'Already active sheet exists'})
        else:
            return JsonResponse({'status': 'No response data'})
       
    except Exception as e:
        # Handle any exceptions that might occur
        return JsonResponse({'status': 'Error', 'error': str(e)})




# def send_values_to_assign_rfid_to_sheet(request, project_id, sheet_id, rfid_value):

#     instance = sheets_rifd.objects.create(project_id = project_id, sheet_id = sheet_id, rfid_value = rfid_value)

#     context = {

#         'project_id' : instance.project_id,
#         'sheet_id' : instance.sheet_id,
#         'status' : 'done'

#     }

#     return JsonResponse(context)


# def assign_rfid_to_sheet(request):

#     instance = values_for_assignment.objects.last()

#     context = {

#         'project_id' : instance.project_id,
#         'sheet_id' : instance.sheet_id,

#     }

#     return JsonResponse(context)



    


def sheet_status_active(request, sheet_id):

    print('----------------')
    product_qr_instance = product_qr.objects.get(id = sheet_id)
    product_qr_instance.status = True
    product_qr_instance.save()

    print(product_qr_instance)

    return redirect(list_generated_product_qr)




def sheet_status_active_demo(request, sheet_id, rfid_value, rfid_reader):

    #check for if values are not equal to 0 

    context = {

        'sheet_id' : sheet_id,
        'rfid_value' : rfid_value,
        'rfid_reader' : rfid_reader,

    }

    return JsonResponse(context)





def sheet_status_deactive(request, sheet_id):

    
    product_qr_instance = product_qr.objects.get(id = sheet_id)
    product_qr_instance.status = False
    product_qr_instance.save()

    print(product_qr_instance)

    return redirect(list_generated_product_qr)





    

    


@login_required(login_url='login')
def download(request):
    # fill these variables with real values


    if request.method == 'POST':

        fl_path =  request.POST.get('link')



        if os.path.exists(fl_path):

            with open(fl_path, 'r' ) as fh:
                mime_type  = mimetypes.guess_type(fl_path)
                print('--------------------')
                print(mime_type)
                response = HttpResponse(fh.read(), content_type=mime_type)
                response['Content-Disposition'] = 'attachment;filename=' + str(fl_path)

                return response



        else:
            messages.error(request, 'path does not exist')


def _delete_file(path):
   """ Deletes file from filesystem. """
   if os.path.isfile(path):
       os.remove(path)

       return 'abc'

@login_required(login_url='login')
def delete_dashboard(request):

    return render(request, 'delete/dashbaord.html')


# delete view




@login_required(login_url='login')
def add_request_material(request):


    if request.method == 'POST':

        forms = request_material_Form(request.POST)

        if forms.is_valid():
            forms.save()

              
            return JsonResponse({'status' : 'done'}, safe=False)


        else:
                
            error = forms.errors.as_json()
            return JsonResponse({'error' : error}, safe=False)

    else:

        forms = request_material_Form()

        context = {
            'form': forms
        }
        return render(request, 'transactions/add_request_material.html', context)



@login_required(login_url='login')
def update_request_material(request, request_material_id ):


    if request.method == 'POST':

        instance_request_material = request_material.objects.get(id = request_material_id)

        forms = request_material_Form(request.POST, instance=instance_request_material)


        if forms.is_valid():


            forms.save()

            return HttpResponseRedirect(reverse('list_request_material'))

               
         
        else:

            instance = request_material.objects.get(id = request_material_id)
            forms = request_material_Form(instance=instance_request_material)

            context = {
                'form': forms,
            }
            

            return render(request, 'transactions/update_request_material.html', context)


    else:

        instance = request_material.objects.get(id = request_material_id)
        forms = request_material_Form(instance = instance)
       

        context = {
            'form': forms,
        }
        return render(request, 'transactions/update_request_material.html', context)


@login_required(login_url='login')
def delete_request_material(request, request_material_id):

    try:
        con = request_material.objects.filter(id = request_material_id).first()
        
        con.delete()

        return HttpResponseRedirect(reverse('list_inward_delete'))


    except:
        return HttpResponseRedirect(reverse('list_inward_delete'))




@login_required(login_url='login')
def list_request_material(request):

  
    data = request_material.objects.all()

    
    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        'data': data,
       
    }

    return render(request, 'transactions/list_request_material.html', context)

@login_required(login_url='login')
def list_project(request):

  
    data = project.objects.all().order_by("-id")

    
    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        'data': data,
        'project_filter': project_filter(),
       
    }

    return render(request, 'transactions/list_projects.html', context)


from django.db.models import Q


@login_required(login_url='login')
def project_report(request):

    code = request.GET.get('item_code')
    customer_id = request.GET.get('customer')
    dc_date_from = request.GET.get('from_DC_date')
    dc_date_to = request.GET.get('to_DC_date')

    print(code)
    print(customer_id)
    print(dc_date_from)
    print(dc_date_to)

    item_code_instance = item_code.objects.get(code = code)

    # Initialize a base queryset
    base_query = project_matarial_production.objects.filter(item_code=item_code_instance)

    if customer_id:
        customer_instance = customer.objects.get(id=customer_id)
        base_query = base_query.filter(project_matarial_qr__project_material__project__customer=customer_instance)

    # Check if dc_date_from and dc_date_to are not None before adding date filters
    if dc_date_from:
        base_query = base_query.filter(project_matarial_qr__project_material__project__DC_date__gte=dc_date_from)

    if dc_date_to:
        base_query = base_query.filter(project_matarial_qr__project_material__project__DC_date__lte=dc_date_to)


    data = base_query.all()
   

    context = {
        'data': data,
        'project_filter': project_filter(),
       
    }

    return render(request, 'transactions/projects_report.html', context)

def download_project_report(request):

    data = project.objects.all()
    project_filters = project_filter(request.GET, queryset=data)
    data = project_filters.qs

    # Create an HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv')
    
    # Set the CSV filename and attachment header
    response['Content-Disposition'] = 'attachment; filename="project_report.csv"'

    # Create a CSV writer and write the header row
    writer = csv.writer(response)
    writer.writerow(['Sr no', 'ITEM CODE', 'QUANTITY' ,'CUSTOMER NAME','DATE'])  # Replace with your actual column names

    # Write the data rows
    for index, item in enumerate(data, start=1):

        for related_obj in item.project_material_re_1.all():

            for project_material_obj in related_obj.project_material_re.all():
                for production_obj in project_material_obj.project_matarial_qr_production.all():
                    
                    writer.writerow([index, production_obj.item_code, production_obj.production_quantity, item.customer, item.DC_date])  # Replace with your actual field names

    return response






from openpyxl import Workbook
from openpyxl.styles import Font
import os



import json

@login_required(login_url='login')
def close_project(request, project_id):

    print('here')

    if request.method == 'POST':

        print('here')

        print(request.POST)


        project_instance = project.objects.get(id = project_id)
        project_instance.completed = True

        
        quantity = request.POST.getlist('quantity[]')
        item_code = request.POST.getlist('item_code[]')
        material = request.POST.getlist('materialsId[]')



        


        return JsonResponse({'status' : 'done'})
    
    else:

        project_instance = project.objects.get(id = project_id)

        forms = project_Form(instance = project_instance)

        data_form = product_Form()

        data = project_material.objects.filter(project = project_instance)

        print(data)

        context = {
            'form': forms,
            'data_form': data_form,
            'data': data,
            'project_id': project_id,
        }

        return render(request, 'transactions/close_project.html', context)


def delete_project(request, project_id):

    project_instance = project.objects.get(id = project_id)

    project_instance.delete()

    return redirect('list_project')

import pusher

from store.forms import *

@login_required(login_url='login')
def add_project(request):


    if request.method == 'POST':

        print(request.POST)

        # Deserialize the JSON data into a Python object
        forms = project_Form(request.POST, request.FILES)
        
        order_id = request.POST.get("order_id")
       

        quantity = request.POST.getlist('production_quantity')
        item_code_id = request.POST.getlist('item_code')
        production_id = request.POST.getlist('production_id')

        print(quantity)
        print(item_code_id)
        print(production_id)


        if forms.is_valid():

            project_instance = forms.save()

            print('valid')


            for a, b, c in zip(production_id, item_code_id, quantity):

                if a and a!= '0':

                    print('here2')

                    project_material_instnace = project_matarial_production.objects.get(id = a)


                    item_code_instance = item_code.objects.get(id = b)

                    project_material_instnace.item_code = item_code_instance
                    project_material_instnace.production_quantity = c
                    
                    project_material_instnace.save()

                    print('here')

                else:

                    print('here3')

                        
                    item_code_instance = item_code.objects.get(id = b)
                    instance = project_matarial_production.objects.create(item_code = item_code_instance, production_quantity = c, project = project_instance)


            a1212 = alert.objects.create(message = "new project created id" + order_id)
            pusher_client = pusher.Pusher(app_id=settings.PUSH_NOTIFICATIONS_SETTINGS["APP_ID"],
                                      key=settings.PUSH_NOTIFICATIONS_SETTINGS["KEY"],
                                      secret=settings.PUSH_NOTIFICATIONS_SETTINGS["SECRET"],
                                      cluster=settings.PUSH_NOTIFICATIONS_SETTINGS["CLUSTER"],
                                      ssl=settings.PUSH_NOTIFICATIONS_SETTINGS["USE_TLS"])

            pusher_client.trigger('alerts', 'new-alert', {'message': a1212.message})

            return redirect('list_project')


        else:
                
            print(forms.errors)
            
            data_form = product_Form()

            context = {
                'form': forms,
                'data_form': data_form,
            }
            return render(request, 'transactions/add_project.html', context)


    else:

        forms = project_Form()

        item_code_data = item_code.objects.all()

        context = {
            'form': forms,
            'item_code_data': item_code_data,
        }
        return render(request, 'transactions/add_project.html', context)

@login_required(login_url='login')
def update_project_accountant(request, project_id):

    project_instance = project.objects.get(id = project_id)

    if request.method == 'POST':

        print(request.POST)

        # Deserialize the JSON data into a Python object
        forms = project_Form(request.POST, instance = project_instance)
        


      
        order_id = request.POST.get("order_id")
       
        quantity = request.POST.getlist('production_quantity')
        item_code_id = request.POST.getlist('item_code')
        production_id = request.POST.getlist('production_id')
        print('-----------------------')
        print(production_id)
        print(quantity)
        print(item_code_id)


        if forms.is_valid():

            print('is valid')

            project_instance = forms.save()


            for a, b, c in zip(production_id, item_code_id, quantity):
                print('----------')
                print(a)
                if a and a != '0':

                    project_material_instnace = project_matarial_production.objects.get(id = a)


                    item_code_instance = item_code.objects.get(id = b)

                    project_material_instnace.item_code = item_code_instance
                    project_material_instnace.production_quantity = c
                    
                    project_material_instnace.save()

                    print('here')

                else:
                        
                    item_code_instance = item_code.objects.get(id = b)
                    instance = project_matarial_production.objects.create(item_code = item_code_instance, project = project_instance, production_quantity = c)

                

            a1212 = alert.objects.create(message = "new project created id" + order_id)
            pusher_client = pusher.Pusher(app_id=settings.PUSH_NOTIFICATIONS_SETTINGS["APP_ID"],
                                      key=settings.PUSH_NOTIFICATIONS_SETTINGS["KEY"],
                                      secret=settings.PUSH_NOTIFICATIONS_SETTINGS["SECRET"],
                                      cluster=settings.PUSH_NOTIFICATIONS_SETTINGS["CLUSTER"],
                                      ssl=settings.PUSH_NOTIFICATIONS_SETTINGS["USE_TLS"])

            pusher_client.trigger('alerts', 'new-alert', {'message': a1212.message})

            return redirect('list_project')


        else:
                
            print(forms.errors)
            
            data_form = product_Form()

            context = {
                'form': forms,
                'data_form': data_form,
                'project_id': project_id,
            }
            return render(request, 'transactions/update_project_accountant.html', context)


    else:

        forms = project_Form(instance=project_instance)

        production_data = project_matarial_production.objects.filter(project = project_instance)

        item_code_data = item_code.objects.all()

        context = {
            'form': forms,
            'production_data': production_data,
            'item_code_data': item_code_data,
            'project_id': project_id,
        }
        return render(request, 'transactions/update_project_accountant.html', context)


def add_project_designer(request, project_id):


    project_instance = project.objects.get(id = project_id)

    if request.method == 'POST':


        print(request.POST)

        forms = project_Form(request.POST, instance=project_instance)
        


        sheet_no_id = request.POST.getlist("no_need")
        order_id = request.POST.get("order_id")
        length = request.POST.getlist("length")
        width = request.POST.getlist("width")

        print('-----------------')

        print(length)
        print(width)

        filtered_values = project_material.objects.filter(project = project_instance).values_list('sheet_no', flat=True)

        existing_values = [int(value) for value in filtered_values]

        # Find values that are unique to value_list
        sheet_no_id = [int(value) for value in sheet_no_id]

        unique_values = [value for value in sheet_no_id if value not in existing_values]

        print(existing_values)
        print(sheet_no_id)
        print(unique_values)

        if forms.is_valid():
            

           


          
            project_instance = forms.save()

            if unique_values != ['']:

                print('in')
                for a, b, c in zip(unique_values, length, width):
                    print('in for')
                    try:

                     

                        aa = product_qr.objects.get(id = a)
                        
                        print('printing a------------------')
                        print(a)

                        print('printing a------------------')
                        obj, created = product.objects.get_or_create(category_id = aa.product.category.id, size_id = aa.product.size.id, grade_id = aa.product.grade.id, thickness_id = aa.product.thickness.id)
                        product_id = obj

                    except product.DoesNotExist:

                        product_id = created

                    project_material_instance = project_material.objects.create(product = product_id, project = project_instance, quantity = 1, sheet_no = a, length = b, width = c)

                    aaaas = project_matarial_qr.objects.create(project_material = project_material_instance)
                    print('-----------------')
                    print(aaaas)
                    print('-----------------')

            a1212 = alert.objects.create(message = "new project created id" + order_id)
            pusher_client = pusher.Pusher(app_id=settings.PUSH_NOTIFICATIONS_SETTINGS["APP_ID"],
                                      key=settings.PUSH_NOTIFICATIONS_SETTINGS["KEY"],
                                      secret=settings.PUSH_NOTIFICATIONS_SETTINGS["SECRET"],
                                      cluster=settings.PUSH_NOTIFICATIONS_SETTINGS["CLUSTER"],
                                      ssl=settings.PUSH_NOTIFICATIONS_SETTINGS["USE_TLS"])

            pusher_client.trigger('alerts', 'new-alert', {'message': a1212.message})

            return redirect('list_project')

        else:

            print(forms.errors)

            data_form = product_Form()

            project_material_data = project_material.objects.filter(project = project_instance)
           
            production_data = project_matarial_production.objects.filter(project = project_instance)



            context = {
                'form': forms,
                'production_data': production_data,
                'project_material_data': project_material_data,
                'data_form': data_form,
                'project_id': project_id,

            }
            return render(request, 'transactions/add_project_designer.html', context)
    
    else:

        forms = project_Form(instance=project_instance)

        data_form = product_Form()

        production_data = project_matarial_production.objects.filter(project = project_instance)
        
        project_material_data = project_material.objects.filter(project = project_instance)
        # project_material_qr = 

        context = {
            'form': forms,
            'production_data': production_data,
            'project_material_data': project_material_data,
            'data_form': data_form,
            'project_id': project_id,
        }
        return render(request, 'transactions/add_project_designer.html', context)



@login_required(login_url='login')
def update_project(request, project_id):

    project_instance = project.objects.get(id = project_id)

    if request.method == 'POST':

        print(request.POST)

        # Deserialize the JSON data into a Python object
        forms = project_Form(request.POST, instance=project_instance)
        sheet_no_id = request.POST.getlist("sheet_no")
        
        for a in sheet_no_id:
            print('in for')
            try:

                aa = product_qr.objects.get(id = a)
                
                print('printing a------------------')
                print(a)

                print('printing a------------------')
                obj, created = product.objects.get_or_create(category_id = aa.product.category.id, size_id = aa.product.size.id, grade_id = aa.product.grade.id, thickness_id = aa.product.thickness.id)
                product_id = obj

            except product.DoesNotExist:

                product_id = created

            project_material_instance = project_material.objects.create(product = product_id, project = project_instance, quantity = 1, sheet_no = a)

            aaaas = project_matarial_qr.objects.create(project_material = project_material_instance)
            print('-----------------')
            print(aaaas)
            print('-----------------')


       
        print(sheet_no_id)

        if forms.is_valid():

            print('is valid')

            project_instance = forms.save()

            
            return redirect('list_project')


        else:
                
            print(forms.errors)
            

            context = {
                'form': forms,
                'data_form': data_form,
            }
            return render(request, 'transactions/update_project.html', context)


    else:

        forms = project_Form(instance=project_instance)

        data_form = product_Form()
        project_material_data = project_material.objects.filter(project = project_instance)
        material_data = project_matarial_qr.objects.filter(project_material__in = project_material_data)

        context = {
            'form': forms,
            'data_form': data_form,
            'material_data': material_data,
            'project_id': project_id,
        }
        return render(request, 'transactions/update_project.html', context)

@login_required(login_url='login')
def get_sheet_details(request):

    sheet_no = request.POST.get('sheet_no')
    print(sheet_no)

    try:

        instance = product_qr.objects.get(id = sheet_no)
    except product_qr.DoesNotExist:

        return JsonResponse({'status' : 'error', 'msg' : 'Sheet No : ' + sheet_no + ' Does not exsist'})
    
    
    data = {
            'size': instance.product.size.id,
            'grade': instance.product.grade.id,
            'thickness': instance.product.thickness.id,
            'category': instance.product.category.id,
            # Add more fields as needed
        }

    print(data)

    return JsonResponse({"status" : 'done', 'data' : data})



@login_required(login_url='login')
def confirm_outward(request, project_id):


    project_instance = project.objects.get(id = project_id)
    

    forms = project_Form(instance = project_instance)
                

    production_data = project_matarial_production.objects.filter(project = project_instance)
    

    context = {
        'form': forms,
        'production_data': production_data,
        'project_id': project_id,
    }
    return render(request, 'transactions/confirm_outward.html', context)



@login_required(login_url='login')
def confirm_small_outward_json(request, project_outward_id):

    instance = project_outward.objects.get(id = project_outward_id)

    instance.date_time = datetime.now()

    instance.save()

    return JsonResponse({"success": True, "message": "Outward confirmed successfully!"}, status = 200)




@login_required(login_url='login')
def submit_invoice_json(request, production_material_id, invoiceId):

    print('------------here-------------')



    print(invoiceId)

    instance = project_matarial_production.objects.get(id = production_material_id)

    instance.invoice_no = invoiceId

    instance.save()

    return redirect('confirm_outward', project_id = instance.project_matarial_production.project.id)






@login_required(login_url='login')
def confirm_main_outward_json(request, project_outward_id):

    instance = project_outward_main_label.objects.get(id = project_outward_id)
    
    instance.date_time = datetime.now()
    instance.save()

    data = project_outward.objects.filter(main_label = instance)
    
    print('-------------------------')
    print('-------------------------')
    print('-------------------------')
    print(data)

    for i in data:


        i.date_time = datetime.now()

        i.save()

    return JsonResponse({"success": True, "message": "Outward confirmed successfully!"}, status = 200)


@login_required(login_url='login')
def add_project_inward(request, project_id):

    project_instance = project.objects.get(id = project_id)
    
    if request.method == 'POST':

        invoice_no = request.POST.getlist("invoice_no")
        title = request.POST.getlist("title")
        quantity = request.POST.getlist("quantity")
        description = request.POST.getlist("description1")
        date = request.POST.getlist("DC_date")



        print(request.POST)
        print(invoice_no)
        print(title)
        print(quantity)
        print(description)
        print(date)


        for a,b,c,d,e,f in zip(invoice_no, title, quantity, description, date):

            project_inward.objects.create(project = project_instance, invoice_no = a, title = b, quantity = c, description = e, date = f)

        
        return redirect('list_project')




def add_project_outward_new(request, production_material_id, small_label, main_label):

    main_label = int(main_label)

    project_matarial_production_instance = project_matarial_production.objects.get(id=production_material_id)
    
    # Clear existing entries
    project_outward.objects.filter(project_matarial_production=project_matarial_production_instance).delete()
    project_outward_main_label.objects.filter(project_matarial_production=project_matarial_production_instance).delete()

    # Update production instance
    project_matarial_production_instance.barcode_count = small_label
    project_matarial_production_instance.main_label_count = main_label
    project_matarial_production_instance.save()

    total_quantity = project_matarial_production_instance.production_quantity

    # Create main label entries
    base_main_quantity = int(total_quantity) // int(main_label)
    main_remainder = int(total_quantity) % int(main_label)

    main_label_entries = []
    for i in range(main_label):
        main_quantity = base_main_quantity + (1 if i < main_remainder else 0)
        main_entry = project_outward_main_label.objects.create(
            project_matarial_production=project_matarial_production_instance, 
            quantity=main_quantity
        )
        main_label_entries.append(main_entry)

    # Distribute small labels within main labels
    small_label = int(small_label)
    small_label_per_main = small_label // main_label
    small_remainder = small_label % main_label

    for idx, main_entry in enumerate(main_label_entries):
        small_count_for_this_main = small_label_per_main + (1 if idx < small_remainder else 0)
        small_base_quantity = main_entry.quantity // small_count_for_this_main
        small_remainder_quantity = main_entry.quantity % small_count_for_this_main

        for i in range(small_count_for_this_main):
            quantity = small_base_quantity + (1 if i < small_remainder_quantity else 0)
            project_outward.objects.create(
                project_matarial_production=project_matarial_production_instance,
                main_label=main_entry,
                quantity=quantity
            )



    return redirect('confirm_outward', project_matarial_production_instance.project.id)
    


    

from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.graphics.barcode import code128
from reportlab.lib.pagesizes import mm


def generate_small_barcode(request, id):
    
    data = project_outward.objects.get(id=id)

    # Create a PDF buffer with 50mm x 50mm page size
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=(50 * mm, 50 * mm))

    # Load the logo
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'RAVIRAJ_LOGO.png')  # Replace with the actual path to your logo
    logo_width = 30 * mm
    logo_height = 9 * mm

    # Center the logo at the top
    logo_x = (50 * mm - logo_width) / 2
    logo_y = 50 * mm - logo_height - 2 * mm  # 2mm margin from the top

    # Draw the logo
    pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')

    # Set the barcode value
    barcode_value = str(data.id)

    # Create the barcode
    barcode = code128.Code128(barcode_value, barWidth=0.5 * mm, barHeight=12 * mm)

    # Center the barcode below the logo
    barcode_x = (50 * mm - barcode.width) / 2
    barcode_y = logo_y - barcode.height - 5 * mm  # Space below the logo

    # Draw the barcode
    barcode.drawOn(pdf, barcode_x, barcode_y)

    # Set text font and size
    pdf.setFont("Helvetica-Bold", 6)  # Adjust font size for better fit

    # Prepare text lines
    text_lines = [
        f"Project ID: {data.project_matarial_production.project.order_id}",
        f"Quantity: {data.quantity}",
        f"Item Code: {data.project_matarial_production.item_code.code}",
        f"Customer Name: {data.project_matarial_production.project.customer.name}",
        f"Supplier Name: Ravi-Raj Anodisers"
    ]

    text_y_start = barcode_y - 5 * mm  # Space below the barcode

    # Reduce line spacing to 3mm
    line_spacing = 3 * mm

    # Center-align each text line
    for line in text_lines:
        text_width = pdf.stringWidth(line, "Helvetica-Bold", 6)  # Adjust font size as needed
        text_x = (50 * mm - text_width) / 2  # Center horizontally
        pdf.drawString(text_x, text_y_start, line)
        text_y_start -= line_spacing  # Minimal spacing between lines

    # Finalize the PDF
    pdf.showPage()
    pdf.save()

    # Get the PDF content
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="barcode_{id}.pdf"'

    return response



def generate_final_barcode(request, project_matarial_production_id):
    
    
    
    data = project_outward_main_label.objects.get(id=project_matarial_production_id)

    # Initialize the PDF buffer and canvas
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=(50 * mm, 50 * mm))

    # Set the logo path
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'RAVIRAJ_LOGO.png')
    logo_width = 30 * mm
    logo_height = 9 * mm

    # Center the logo at the top
    logo_x = (50 * mm - logo_width) / 2
    logo_y = 50 * mm - logo_height - 2 * mm  # 2mm margin from the top

    # Draw the logo
    pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')

    # Set the barcode value
    barcode_value = str(data.id)

    # Create the barcode
    barcode = code128.Code128(barcode_value, barWidth=0.5 * mm, barHeight=12 * mm)

    # Center the barcode below the logo
    barcode_x = (50 * mm - barcode.width) / 2
    barcode_y = logo_y - barcode.height - 5 * mm  # Space below the logo

    # Draw the barcode
    barcode.drawOn(pdf, barcode_x, barcode_y)

    # Set text font and size
    pdf.setFont("Helvetica-Bold", 6)  # Adjust font size for better fit

    # Prepare text lines
    text_lines = [
        f"Project ID: {data.project_matarial_production.project.order_id}",
        f"Quantity: {data.quantity}",
        f"Item Code: {data.project_matarial_production.item_code.code}",
        f"Customer Name: {data.project_matarial_production.project.customer.name}",
        f"Supplier Name: Ravi-Raj Anodisers"
    ]

    # Space below the barcode for text
    text_y_start = barcode_y - 5 * mm
    line_spacing = 3 * mm  # Reduce line spacing to 3mm

    # Center-align each text line
    for line in text_lines:
        text_width = pdf.stringWidth(line, "Helvetica-Bold", 6)
        text_x = (50 * mm - text_width) / 2
        pdf.drawString(text_x, text_y_start, line)
        text_y_start -= line_spacing

    # Add a small-font note below the main text
    note = "* This is main box label with total itemcode quantity"
    pdf.setFont("Helvetica", 4)  # Smaller font size
    note_width = pdf.stringWidth(note, "Helvetica", 4)
    note_x = (50 * mm - note_width) / 2
    note_y = 3 * mm  # 2mm spacing below the last line of text
    pdf.drawString(note_x, note_y, note)

    # Finalize the PDF
    pdf.showPage()
    pdf.save()

    # Get the PDF content
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="main_box_label_{data.id}.pdf"'

    return response

def generate_all_barcode(request, project_matarial_production_id):

    project_material = project_matarial_production.objects.get(id=project_matarial_production_id)
    data = project_outward_main_label.objects.filter(project_matarial_production=project_material).select_related('project_matarial_production__project', 'project_matarial_production__item_code')


    if not data.exists():
        return HttpResponse("No data available to generate barcodes.", status=400)

    # Initialize the PDF buffer and canvas
    pdf_buffer = BytesIO()
    pdf = canvas.Canvas(pdf_buffer, pagesize=(50 * mm, 50 * mm))

    # Set the logo path
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'RAVIRAJ_LOGO.png')
    if not os.path.exists(logo_path):
        return HttpResponse("Logo not found.", status=404)

    # Set the logo path
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'RAVIRAJ_LOGO.png')
    logo_width = 30 * mm
    logo_height = 9 * mm

    # for main label
    for i in data:
        
        # Center the logo at the top
        logo_x = (50 * mm - logo_width) / 2
        logo_y = 50 * mm - logo_height - 2 * mm  # 2mm margin from the top

        pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')

        barcode_value = str(i.id)

        barcode = code128.Code128(barcode_value, barWidth=0.5 * mm, barHeight=12 * mm)

        
        barcode_x = (50 * mm - barcode.width) / 2
        barcode_y = logo_y - barcode.height - 5 * mm 

        barcode.drawOn(pdf, barcode_x, barcode_y)

        pdf.setFont("Helvetica-Bold", 6)  

        text_lines = [
            f"Project ID: {i.project_matarial_production.project.order_id}",
            f"Quantity: {i.quantity}",
            f"Item Code: {i.project_matarial_production.item_code.code}",
            f"Customer Name: {i.project_matarial_production.project.customer.name}",
            f"Supplier Name: Ravi-Raj Anodisers"
        ]

        text_y_start = barcode_y - 5 * mm
        line_spacing = 3 * mm  

        for line in text_lines:
            text_width = pdf.stringWidth(line, "Helvetica-Bold", 6)
            text_x = (50 * mm - text_width) / 2
            pdf.drawString(text_x, text_y_start, line)
            text_y_start -= line_spacing

        note = "* This is main box label with total itemcode quantity"
        pdf.setFont("Helvetica", 4)  # Smaller font size
        note_width = pdf.stringWidth(note, "Helvetica", 4)
        note_x = (50 * mm - note_width) / 2
        note_y = 3 * mm  # 2mm spacing below the last line of text
        pdf.drawString(note_x, note_y, note)

        # Finalize the page and move to the next
        pdf.showPage()


    # for small labels
    for i in data:

        small_label_data = i.main_label_re.all()

        for z in small_label_data:
        
            # Center the logo at the top
            logo_x = (50 * mm - logo_width) / 2
            logo_y = 50 * mm - logo_height - 2 * mm  # 2mm margin from the top

            pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')

            barcode_value = str(z.id)

            barcode = code128.Code128(barcode_value, barWidth=0.5 * mm, barHeight=12 * mm)

            
            barcode_x = (50 * mm - barcode.width) / 2
            barcode_y = logo_y - barcode.height - 5 * mm 

            barcode.drawOn(pdf, barcode_x, barcode_y)

            pdf.setFont("Helvetica-Bold", 6)  

            text_lines = [
                f"Project ID: {z.project_matarial_production.project.order_id}",
                f"Quantity: {z.quantity}",
                f"Item Code: {z.project_matarial_production.item_code.code}",
                f"Customer Name: {z.project_matarial_production.project.customer.name}",
                f"Supplier Name: Ravi-Raj Anodisers"
            ]

            text_y_start = barcode_y - 5 * mm
            line_spacing = 3 * mm  

            for line in text_lines:
                text_width = pdf.stringWidth(line, "Helvetica-Bold", 6)
                text_x = (50 * mm - text_width) / 2
                pdf.drawString(text_x, text_y_start, line)
                text_y_start -= line_spacing

            # Finalize the page and move to the next
            pdf.showPage()

    
    pdf.save()

    # Get the PDF content
    pdf_buffer.seek(0)
    response = HttpResponse(pdf_buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="main_box_label.pdf"'

    return response

    

# def generate_all_final_barcode(request, project_matarial_production_id):
    
    
    
#     data = project_matarial_production.objects.get(id=project_matarial_production_id)
#     ids_list = project_outward_main_label.objects.filter(project_matarial_production=data)

#     # Initialize the PDF buffer and canvas
#     pdf_buffer = BytesIO()
#     pdf = canvas.Canvas(pdf_buffer, pagesize=(50 * mm, 50 * mm))

#     # Set the logo path
#     logo_path = os.path.join(settings.BASE_DIR, 'static', 'RAVIRAJ_LOGO.png')
#     logo_width = 30 * mm
#     logo_height = 9 * mm

#     # Center the logo at the top
#     logo_x = (50 * mm - logo_width) / 2
#     logo_y = 50 * mm - logo_height - 2 * mm  # 2mm margin from the top

#      # Loop through each record and create a separate page for each sticker
#     for id in ids_list:
#         # Draw the logo
#         pdf.drawImage(logo_path, logo_x, logo_y, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')

#         # Generate the barcode
#         barcode_value = str(id.id)  # Ensure the barcode value is valid
#         barcode = code128.Code128(barcode_value, barWidth=0.5 * mm, barHeight=12 * mm)
#         barcode_x = (50 * mm - barcode.width) / 2
#         barcode_y = logo_y - barcode.height - 5 * mm  # Space below the logo
#         barcode.drawOn(pdf, barcode_x, barcode_y)

#         # Add text below the barcode
#         pdf.setFont("Helvetica-Bold", 6)  # Font size for text
#         text_lines = [
#             f"Project ID: {id.project_matarial_production.project.order_id}",
#             f"Quantity: {id.quantity}",
#             f"Item Code: {id.project_matarial_production.item_code.code}",
#             f"Customer Name: {id.project_matarial_production.project.customer.name}",
#             f"Supplier Name: Ravi-Raj Anodisers"
#         ]
#         text_y_start = barcode_y - 5 * mm  # Space below the barcode
#         line_spacing = 3 * mm

#         for line in text_lines:
#             text_width = pdf.stringWidth(line, "Helvetica-Bold", 6)
#             text_x = (50 * mm - text_width) / 2  # Center-align text
#             pdf.drawString(text_x, text_y_start, line)
#             text_y_start -= line_spacing  # Decrease y-coordinate for next line

   
#         # Add a small-font note below the main text
#         note = "* This is main box label with total itemcode quantity"
#         pdf.setFont("Helvetica", 4)  # Smaller font size
#         note_width = pdf.stringWidth(note, "Helvetica", 4)
#         note_x = (50 * mm - note_width) / 2
#         note_y = 3 * mm  # 2mm spacing below the last line of text
#         pdf.drawString(note_x, note_y, note)

#         # Finalize the PDF
#         pdf.showPage()
        
#     pdf.save()

#     # Get the PDF content
#     pdf_buffer.seek(0)
#     response = HttpResponse(pdf_buffer, content_type='application/pdf')
#     response['Content-Disposition'] = f'attachment; filename="main_box_label_{data.id}.pdf"'

#     return response



def scan_barcode(request):
    if request.method == 'POST':
        barcode = request.POST.get('barcode')
        print(barcode)

        data = project_outward.objects.get(id = barcode)
        
        try:
            # Look up the product by the scanned barcode

            return redirect('confirm_outward', project_id=data.project_matarial_production.project.id)

        
        except project_outward.DoesNotExist:
            return HttpResponse("Product not found.")
    
    return render(request, 'transactions/scan_barcode.html')  # Return scan page if no POST data


def scan_main_barcode(request):

    if request.method == 'POST':

        barcode = request.POST.get('barcode')
        print(barcode)

        data = project_outward_main_label.objects.get(id = barcode)
        
        try:
            # Look up the product by the scanned barcode

            return redirect('confirm_outward', project_id=data.project_matarial_production.project.id)

        
        except project_outward.DoesNotExist:
            return HttpResponse("Product not found.")
    

    else:
       
    
        return render(request, 'transactions/scan_main_barcode.html')  # Return scan page if no POST data




@login_required(login_url='login')
def add_inward_item_code(request):

    if request.method == 'POST':

        forms = inward_item_code_Form(request.POST)

        if forms.is_valid():
            forms.save()
            return redirect('list_inward_item_code')
        else:

            context = {
                'form': forms
            }
            return render(request, 'transactions/add_inward_item_code.html', context)

    
    else:

        forms = inward_item_code_Form()

        context = {
            'form': forms
        }
        return render(request, 'transactions/add_inward_item_code.html', context)

        

@login_required(login_url='login')
def update_inward_item_code(request, inward_item_code_id):

    if request.method == 'POST':

        instance = inward_item_code.objects.get(id=inward_item_code_id)

        forms = inward_item_code_Form(request.POST, instance=instance)

        if forms.is_valid():
            forms.save()
            return redirect('list_inward_item_code')
        else:
            print(forms.errors)
    
    else:

        instance = inward_item_code.objects.get(id=inward_item_code_id)
        forms = inward_item_code_Form(instance=instance)

        context = {
            'form': forms
        }
        return render(request, 'transactions/add_inward_item_code.html', context)

        

@login_required(login_url='login')
def delete_inward_item_code(request, inward_item_code_id):

    inward_item_code.objects.get(id=inward_item_code_id).delete()

    return HttpResponseRedirect(reverse('list_inward_item_code'))


@login_required(login_url='login')
def list_inward_item_code(request):

    data = inward_item_code.objects.all()  

    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)


    context = {
        'data': data
    }
    return render(request, 'transactions/list_inward_item_code.html', context)





@login_required(login_url='login')
def add_inward(request):


    if request.method == 'POST':

        
        # Deserialize the JSON data into a Python object
        forms = project_inward_Form(request.POST, request.FILES)
        

        if forms.is_valid():

            forms.save()

            return redirect('list_inward')


        else:
                
            print(forms.errors)
            
            data_form = project_inward_Form()

            item_code_data = inward_item_code.objects.all()

            context = {
                'form': forms,
                'data_form': data_form,
                'item_code_data': item_code_data,
            }
            return render(request, 'transactions/add_inward.html', context)


    else:

        forms = project_inward_Form()
        item_code_data = inward_item_code.objects.all()

         
        context = {
            'form': forms,
                'item_code_data': item_code_data,

        }
        return render(request, 'transactions/add_inward.html', context)
    


@login_required(login_url='login')

def inward_itemcode_description(request):
    item_code_id = request.GET.get('item_code_id')
    try:
        item = inward_item_code.objects.get(pk=item_code_id)
        return JsonResponse({
            'description': item.description,
        })
    except inward_item_code.DoesNotExist:
        return JsonResponse({'error': 'Item not found'}, status=404)


@login_required(login_url='login')
def update_inward(request, inward_id):

    instance = project_inward.objects.get(id = inward_id)


    if request.method == 'POST':

        
        # Deserialize the JSON data into a Python object
        forms = project_inward_Form(request.POST, instance=instance)
        

        if forms.is_valid():

            forms.save()

            return redirect('list_inward')


        else:
                
            print(forms.errors)
            
            data_form = project_inward_Form(instance=instance)

            context = {
                'form': forms,
                'data_form': data_form,
            }
            return render(request, 'transactions/add_inward.html', context)


    else:

        forms = project_inward_Form(instance=instance)

        context = {
            'form': forms,
        }
        return render(request, 'transactions/add_inward.html', context)
    



@login_required(login_url='login')
def list_inward(request):

    data = project_inward.objects.all().order_by('-id')

    data = project_inward_filter(request.GET, queryset=data)

    data = data.qs

    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        'data': data,
        'project_filter': project_inward_filter(request.GET),
       
    }
    


    return render(request, 'transactions/list_inward.html', context)



@login_required(login_url='login')
def delete_inward(request, inward_id):

    project_inward.objects.get(id = inward_id).delete()


    return redirect('list_inward')



@login_required(login_url='login')
def inward_report(request):


    data = project_inward.objects.all().order_by("-project")

    
    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        'data': data,
        'project_inward_filter': project_inward_filter(),
       
    }

    return render(request, 'transactions/list_project_inward_report.html', context)




def download_inward_report(request):

      
    data = project_inward.objects.all()
    project_inward_filters = project_inward_filter(request.GET, queryset=data)
    project_inward_filters_data1 = list(project_inward_filters.qs.values_list('customer', 'quantity', 'description', 'date'))
    project_inward_filters_data = list(map(list, project_inward_filters_data1))

    # Create an Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inward Report'

    # Header information
    ws.append([''])  # Empty row
    ws.append(['Inward Report'])  # Title row
    ws.append([''])  # Empty row
    ws.append([''])  # Empty row

    # Add table headers
    headers = ["Sr No", "Customer", "Quantity", "Description", "Date"]
    ws.append(headers)

    # Apply light green color to header row
    light_green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for cell in ws[ws.max_row]:  # Select the last added row (header row)
        cell.fill = light_green_fill

    # Append data to the Excel file
    counteer = 1
    for i in project_inward_filters_data:
        ws.append([counteer, i[0], i[1], i[2], i[3]])
        counteer += 1

    # Save the workbook to a BytesIO stream
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create a response for the file download
    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Inward_Report.xlsx"'

    return response


def download_inward_report_pdf(request):

      
    data = project_inward.objects.all()
    project_inward_filters = project_inward_filter(request.GET, queryset=data)
    project_inward_filters_data1 = list(project_inward_filters.qs.values_list('inward_supplier', 'quantity', 'inward_item_code__description', 'date'))
    project_inward_filters_data = list(map(list, project_inward_filters_data1))

    # Add serial numbers to the data
    formatted_data = [[idx + 1] + row for idx, row in enumerate(project_inward_filters_data)]

    # Prepare the table data with headers
    table_data = [['Sr No', 'Customer', 'Quantity', 'Description', 'Date']]  # Header row
    table_data.extend(formatted_data)

    # Set up the PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    # Define the styles for the heading
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Create the title
    title = Paragraph("Inward Report", title_style)

    # Spacer between the title and table
    spacer = Spacer(1, 20)

    # Create the table
    table = Table(table_data)

    # Add styling to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),  # Header row background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all text
        ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),  # Data row background
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Borders for table cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold font for header
    ]))

    # Build the PDF
    elements = [title, spacer, table]
    doc.build(elements)

    # Prepare PDF as response
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf', headers={
        'Content-Disposition': 'attachment; filename="inward_report.pdf"'
    })


from django.utils import timezone

from collections import defaultdict


from collections import defaultdict
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
from django.utils import timezone
from django.conf import settings
from django.core.mail import EmailMessage
from django.http import HttpResponse
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph

def draw_border(canvas, doc):
    width, height = A4  # Get the width and height of the page
    canvas.saveState()
    canvas.setStrokeColor(colors.black)  # Set border color
    canvas.setLineWidth(1)  # Set border width
    canvas.rect(10, 10, width-20, height-20)  # Draw the rectangle border
    canvas.restoreState()

def email_inward_report(request):
    today = timezone.now().date()
    
    # Filter project_outward entries for today
    outward_today = project_inward.objects.filter(date = today)

    # Collect related project instances
    
    # Prepare data for the PDF
    buffer = BytesIO()
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    
    elements = []
    styles = getSampleStyleSheet()

    # Modify the 'Title' style for the main title
    title_style = styles['Title']
    title_style.alignment = 1
    title_style.fontSize = 14

    # Create a new style for the secondary title
    title_style2 = ParagraphStyle(
        'Title2',
        parent=styles['Title'],
        alignment=1,
        fontSize=12
    )

    # Title
    title = Paragraph("Ravi Raj Anodisers", title_style)
    elements.append(title)

    # Secondary Title
    title2 = Paragraph("Inward Report", title_style2)
    elements.append(title2)

    counteer = 1
    # Loop over each related project
    for i in outward_today:
    # Filter outward entries specific to this project
      

        # Main table for project details
        main_table_data = [
            ["#", "Item Code", "Item Code Description", "Supplier", "Quantity", "Description", "Date"],
            [   
                str(counteer),
                str(i.inward_item_code.item_code),
                str(i.inward_item_code.description),
                str(i.inward_supplier),
                str(i.quantity),
                str(i.description),
                str(i.date),
            ]
        ]

        counteer += 1

        # Create the main table
        main_table = Table(main_table_data, colWidths=[30, 50, 120, 100, 50, 150, 50])
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 8),  # Set font size for all cells
        ]))

        elements.append(main_table)
        elements.append(Spacer(1, 0.5 * cm))

       
    # Build and save the PDF
    pdf.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)
    file_path = os.path.join(settings.MEDIA_ROOT, 'Inward_report.pdf')
    with open(file_path, 'wb') as f:
        f.write(buffer.getvalue())
    buffer.close()

    # Create the email
    email = EmailMessage(
        subject='Inward Report PDF',
        body='Please find the attached Inward Report in PDF format.',
        from_email='rradailyupdates@gmail.com',
        to=['varad@ravirajanodisers.com', 'ravi@ravirajanodisers.com', 'pratikgosavi654@gmail.com', 'raj@ravirajanodisers.com'],
        # to=['pratikgosavi654@gmail.com'],
    )

    # Attach the generated PDF
    email.attach_file(file_path)

    # Send the email
    email.send()

    return JsonResponse({'message': 'response message'}, status=400)

def confrim_outward_report(request):
    today = timezone.now().date()
    
    # Filter project_outward entries for today
    outward_today = project_outward.objects.filter(date_time__date=today)

    # Collect related project instances
    related_projects = {outward.project_matarial_production.project for outward in outward_today}
    
    # Prepare data for the PDF
    buffer = BytesIO()
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1

    # Title
    title = Paragraph("Material Outward Report", title_style)
    elements.append(title)

    light_orange = colors.Color(1, 0.647, 0, alpha=1)  # RGB for light orange

    # Loop over each related project
    for project_instance in related_projects:
        # Filter outward entries specific to this project
        project_outward_entries = outward_today.filter(project_matarial_production__project=project_instance)

        # Aggregate quantity by item code for this project
        item_code_totals = defaultdict(int)
        for outward in project_outward_entries:
            item_code = outward.project_matarial_production.item_code
            item_code_totals[item_code] += outward.quantity

        # Main table for project details
        main_table_data = [
            ["#", "Date", "Project Order No", "RRA Invoice No", "Customer Name"],
            [
                str(project_instance.id),
                str(project_instance.DC_date),
                str(project_instance.order_id),
                str(project_instance.rra_invoice_no),
                str(project_instance.customer),
            ]
        ]

        # Create the main table
        main_table = Table(main_table_data)
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(main_table)
        elements.append(Spacer(1, 0.5 * cm))

        # Nested table for item codes and quantities
        nested_table_data = [["#", "Item Code", "Quantity"]]
        for counter, (item_code, total_quantity) in enumerate(item_code_totals.items(), start=1):
            nested_table_data.append([
                str(counter),
                str(item_code),
                str(total_quantity),
            ])

        nested_table = Table(nested_table_data, colWidths=[40, 400, 100])
        nested_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light_orange),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))

        elements.append(nested_table)
        elements.append(Spacer(1, 0.5 * cm))

    # Build and save the PDF
    pdf.build(elements, onFirstPage=draw_border, onLaterPages=draw_border)

    file_path = os.path.join(settings.MEDIA_ROOT, 'outward_report.pdf')
    with open(file_path, 'wb') as f:
        f.write(buffer.getvalue())
    buffer.close()

    # Create the email
    email = EmailMessage(
        subject='Outward Report PDF',
        body='Please find the attached outward report in PDF format.',
        from_email='rradailyupdates@gmail.com',
        to=['varad@ravirajanodisers.com', 'ravi@ravirajanodisers.com', 'pratikgosavi654@gmail.com', 'raj@ravirajanodisers.com'],
        # to=['pratikgosavi654@gmail.com'],
    )

    # Attach the generated PDF
    email.attach_file(file_path)

    # Send the email
    email.send()

    return JsonResponse({'message': 'response message'}, status=400)

    # Apply production filter and prefetch the filtered production data
    # data = project_filter_data  # Ensure this is a queryset

    
    # page = request.GET.get('page', 1)
    # paginator = Paginator(data, 50)

    # try:
    #     data = paginator.page(page)
    # except PageNotAnInteger:
    #     data = paginator.page(1)
    # except EmptyPage:
    #     data = paginator.page(paginator.num_pages)




    # context = {
    #     'data': data,
    #     'project_filter': ProjectOutwardMainLabelFilter(request.GET),
       
    # }

      

   


    # return render(request, 'transactions/list_confrim_project_outward_report.html', context)



@login_required(login_url='login')
def outward_report(request):


    projects = project.objects.all().order_by("-id")

     # Check if request.GET is empty
    if is_invalid_get_params(request.GET):
        # Create a mutable copy of request.GET
        default_params = QueryDict(mutable=True)
        
        # Set default parameters

        default_params.update({
        'from_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0) - timedelta(days=1),
        'to_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)
        })

    else:

        from_date_time = unquote(request.GET.get('from_date_time'))
        to_date_time = unquote(request.GET.get('to_date_time'))

        try:
            # Parse the datetime strings into datetime objects
            from_date_time_obj = datetime.strptime(from_date_time, '%Y-%m-%dT%H:%M')
            to_date_time_obj = datetime.strptime(to_date_time, '%Y-%m-%dT%H:%M')
        except ValueError:
            return HttpResponse("Invalid datetime format", status=400)

        
    default_params = request.GET.copy()  

   
    default_params.update({
        'from_main_label_date': from_date_time_obj.isoformat(),
        'to_main_label_date': to_date_time_obj.isoformat()
    })


    print('--------------------------------------')
    print(default_params)
    print('--------------------------------------')
            
    project_filter_data = project_filter(default_params, queryset=project.objects.all())
    production_filter = ProjectOutwardMainLabelFilter(default_params, queryset=project_matarial_production.objects.all())


    print(default_params)

    # Apply production filter and prefetch the filtered production data
    project_qs = project_filter_data.qs  # Ensure this is a queryset
    production_qs = production_filter.qs  # Ensure this is a queryset

    print('production_qs')
    print(production_qs)
    print('project_qs')
    print(project_qs)

    # Prefetch the production data using the correct related_name for the relationship
    data = project_qs.filter(
        project_production_n__in=production_qs  # Ensures only projects with related filtered production records
    ).prefetch_related(
        Prefetch('project_production_n', queryset=production_qs)  # Prefetch the filtered production data
    ).distinct()

    
    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)


    for i in data:
        totals = i.project_production_n.aggregate(
            total_quantity=Sum('production_quantity'),
        )
        
        # Add totals to i instance for template access
        i.total_quantity = totals['total_quantity'] or 0


    context = {
        'data': data,
        'project_filter': project_filter(request.GET),
        'project_matarial_production_filter': project_matarial_production_filter(request.GET),
       
    }


    return render(request, 'transactions/list_project_outward_report.html', context)



from openpyxl.styles import PatternFill, Font, Border, Side
import os
import mimetypes
from django.http import HttpResponse
from django.db.models import Sum



from django.core.mail import EmailMessage


from urllib.parse import unquote

# Decode the datetime string


def outward_report_csv(request):

# Query data
    projects = project.objects.all().order_by("-id")

     # Check if request.GET is empty
    if is_invalid_get_params(request.GET):
        # Create a mutable copy of request.GET
        default_params = QueryDict(mutable=True)
        
        # Set default parameters

        default_params.update({
        'from_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0) - timedelta(days=1),
        'to_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)
        })

    else:

        from_date_time = unquote(request.GET.get('from_date_time'))
        to_date_time = unquote(request.GET.get('to_date_time'))

        try:
            # Parse the datetime strings into datetime objects
            from_date_time_obj = datetime.strptime(from_date_time, '%Y-%m-%dT%H:%M')
            to_date_time_obj = datetime.strptime(to_date_time, '%Y-%m-%dT%H:%M')
        except ValueError:
            return HttpResponse("Invalid datetime format", status=400)

        # You can use the parsed datetime objects now
        # For example, filtering or other operations based on the parsed datetime
        # Assuming you may want to update default_params with the valid parsed times
        default_params = request.GET.copy()  # Use a mutable copy if you want to modify it

        # Optional: You can update `default_params` with the parsed values if needed
        default_params.update({
            'from_date_time': from_date_time_obj.strftime('%Y-%m-%dT%H:%M'),
            'to_date_time': to_date_time_obj.strftime('%Y-%m-%dT%H:%M')
        })

                
    project_filter_data = project_filter(default_params, queryset=project.objects.all())
    production_filter = project_matarial_production_filter(default_params, queryset=project_matarial_production.objects.all())


    print(default_params)

    # Apply production filter and prefetch the filtered production data
    project_qs = project_filter_data.qs  # Ensure this is a queryset
    production_qs = production_filter.qs  # Ensure this is a queryset

    print(production_qs)
    print(project_qs)

    # Prefetch the production data using the correct related_name for the relationship
    projects = project_qs.filter(
        project_production_n__in=production_qs  # Ensures only projects with related filtered production records
    ).prefetch_related(
        Prefetch('project_production_n', queryset=production_qs)  # Prefetch the filtered production data
    ).distinct()

    data = projects

    # Create a workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define styles
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    light_orange_fill = PatternFill(start_color='FFDAB9', end_color='FFDAB9', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    bold_font = Font(bold=True)

    # Apply borders to a range of cells in bulk
    def apply_borders_range(start_row, end_row):
        for row in ws.iter_rows(min_row=start_row, max_row=end_row):
            for cell in row:
                if cell.value:  # Only apply borders to non-empty cells
                    cell.border = thin_border

    # Write the main table header
    ws.append(["Material outward", "", "", "", ""])
    ws.append(["#", "Date", "Project Order No", "RRA Invoice No", "Customer Name"])

    # Apply background color to header cells of the main table
    for cell in ws[2]:
        cell.fill = light_green_fill

    row_index = 3

    # Write rows for each project
    for project_instance in data:
        ws.append([])  # Blank row for spacing

        # Write the project row
        ws.append([
            str(project_instance.pk),
            str(project_instance.DC_date),
            str(project_instance.order_id),
            str(project_instance.rra_invoice_no),
            str(project_instance.customer)
        ])

        # Write nested table header for project productions
        ws.append(["", "Item Code", "Quantity"])
        for cell in ws[ws.max_row]:
            cell.fill = light_orange_fill

        # Write project production rows
        productions = project_instance.project_production_n.all()
        for i in productions:
            ws.append([
                "", 
                str(i.item_code), 
                str(i.production_quantity), 
            ])

        # Calculate totals
        totals = project_instance.project_production_n.aggregate(
            total_quantity=Sum('production_quantity'),
        )
        total_quantity = totals['total_quantity'] or 0

        # Write the totals row
        ws.append([
            "", 
            "Project Total", 
            str(total_quantity), 
        ])
        for cell in ws[ws.max_row]:
            cell.font = bold_font

        # Apply borders only after writing all rows
        apply_borders_range(row_index, ws.max_row)

        # Update the row index
        row_index = ws.max_row + 2

    # Save and return the workbook
    name = 'outward_report_with_optimized_colors.xlsx'
    link = os.path.join(BASE_DIR, 'static', 'csv', name)
    wb.save(link)

    mime_type, _ = mimetypes.guess_type(link)
    if mime_type is None:
        mime_type = 'application/octet-stream'

    # with open(link, 'rb') as f:
    #     response = HttpResponse(f.read(), content_type=mime_type)
    #     response['Content-Disposition'] = f'attachment; filename="{name}"'

     # Save the workbook to an in-memory file (BytesIO)
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create the response
    response = HttpResponse(
        excel_file.getvalue(), 
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="outward_report.xlsx"'
    
    return response



from io import BytesIO
from django.conf import settings
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.shortcuts import render
from .models import project
from django.db.models import Sum
import os
from reportlab.lib.units import cm

from django.http import QueryDict



def is_invalid_get_params(get_params):
    # Check if 'undefined' exists and has no value, or if the GET request is empty
    return not get_params or ('undefined' in get_params and len(get_params['undefined']) == 0)



from django.db.models import Prefetch


def generate_outward_report_pdf(request):

     # Filter projects and production based on request parameters
    project_filter = project_filter(request.GET, queryset=project.objects.all())
    production_filter = project_matarial_production_filter(request.GET, queryset=project_matarial_production.objects.filter(date_time__gte=yesterday_6pm, date_time__lte=today_6pm))
    
    # Apply production filter and prefetch the filtered production data
    projects = project_filter.qs.prefetch_related(
        Prefetch('project_production_n', 
                 queryset=production_filter.qs)
    ).distinct()

    context = {
        'projects': projects,
        'project_filter': project_filter,
        'production_filter': production_filter
    }
    
    return render(request, 'report_template.html', context)


def generate_outward_report_daily_pdf(request):
    # Create an in-memory buffer for the PDF file
    
   
    # Create an in-memory buffer for the PDF file
    buffer = BytesIO()

    # Create a PDF document with minimal margins
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1  # Centered

    # Title of the PDF
    title = Paragraph("Material Outward Report", title_style)
    elements.append(title)

    # Query data from the database
    projects = project.objects.all().order_by("-id")

     # Check if request.GET is empty
        # Create a mutable copy of request.GET
        
        # Set default parameters
                
    yesterday_6pm = datetime.now().replace(hour=18, minute=0, second=0, microsecond=0) - timedelta(days=1)
    today_6pm = datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)


    # Prefetch production data related to projects with filtering
    projects = project.objects.prefetch_related(
        Prefetch('project_production_n', 
                queryset=project_matarial_production.objects.filter(date_time__gte=yesterday_6pm, date_time__lte=today_6pm))
    ).filter(project_production_n__date_time__gte=yesterday_6pm, project_production_n__date_time__lte=today_6pm).distinct()

    print(projects)

      
        # Update the parameters
      


    light_orange = colors.Color(1, 0.647, 0, alpha=1)  # RGB for light orange

    for project_instance in projects:
        # Data for the main table
        main_table_data = [
            ["#", "Date", "Project Order No", "RRA Invoice No", "Customer Name"],
            [
                str(project_instance.pk),
                str(project_instance.DC_date),
                str(project_instance.order_id),
                str(project_instance.rra_invoice_no),
                str(project_instance.customer),
            ]
        ]

        # Create the main table
        main_table = Table(main_table_data)
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),  # Header row background
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),  # Data row background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Borders for table cells
        ]))

        elements.append(main_table)

        # Add an empty line for separation
        elements.append(Paragraph(''))

        # Nested table data
        nested_table_data = [
            ["#", "Item Code", "Quantity"]
        ]

        # Add data rows for each production
        productions = project_instance.project_production_n.all()
        for counter, production in enumerate(productions, start=1):
            nested_table_data.append([
                str(counter),
                str(production.item_code),
                str(production.production_quantity),
            ])

        # Add totals for this project
        totals = project_instance.project_production_n.aggregate(
            total_quantity=Sum('production_quantity'),
        )
        total_quantity = totals['total_quantity'] or 0
        nested_table_data.append([
            "",
            "Project Total",
            str(total_quantity),
        ])

        # Create the nested table
        nested_table = Table(nested_table_data, colWidths=[40, 200, 100, 100])  # Adjusted width for Item Code
        nested_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light_orange),  # Header row background
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),  # Data row background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Borders for table cells
        ]))

        # Center the nested table within the page
        elements.append(Paragraph(''))  # Add a blank line for separation
        elements.append(nested_table)

        # Add an extra empty line for spacing between projects
        elements.append(Paragraph(''))


        # Add a blank line
        elements.append(Spacer(1, 0.5 * cm))



    # Build the PDF into the buffer
    pdf.build(elements)

    file_path = os.path.join(settings.MEDIA_ROOT, 'outward_report.pdf')


    with open(file_path, 'wb') as f:
        f.write(buffer.getvalue())

    # Close the buffer
    buffer.close()

    return file_path




# Function to send the PDF as an email attachment
from django.core.mail import EmailMessage
from django.core.mail import EmailMessage
from django.http import HttpResponse

from django.http import HttpResponse
from django.conf import settings
import os




def generate_inward_report_daily_pdf(request):


    default_params = QueryDict(mutable=True)
    
    # Set default parameters

    default_params.update({
    'from_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0) - timedelta(days=1),
    'to_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)
    })

    data = project_inward.objects.all()
    project_inward_filters = project_inward_filter(default_params, queryset=data)
    project_inward_filters_data1 = list(project_inward_filters.qs.values_list('customer', 'quantity', 'description', 'date'))
    project_inward_filters_data = list(map(list, project_inward_filters_data1))

    # Add serial numbers to the data
    formatted_data = [[idx + 1] + row for idx, row in enumerate(project_inward_filters_data)]

    # Prepare the table data with headers
    table_data = [['Sr No', 'Customer', 'Quantity', 'Description', 'Date']]  # Header row
    table_data.extend(formatted_data)

    # Set up the PDF file path
    name = "inward_report.pdf"
    file_path = os.path.join(settings.BASE_DIR, 'static', 'reports', name)  # Ensure the 'static/reports/' directory exists

    # Create the PDF document and save it to the specified path
    doc = SimpleDocTemplate(file_path, pagesize=letter)

    # Define the styles for the heading
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    
    # Create the title
    title = Paragraph("Inward Report", title_style)

    # Spacer between the title and table
    spacer = Spacer(1, 20)

    # Create the table
    table = Table(table_data)

    # Add styling to the table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),  # Header row background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all text
        ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),  # Data row background
        ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Borders for table cells
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),  # Bold font for header
    ]))

    # Build the PDF
    elements = [title, spacer, table]
    doc.build(elements)

    # Return the file path
    return file_path

def send_outward_report_pdf_email_daily(request):
    # Generate and save the PDF
    
    
    file_path = generate_outward_report_daily_pdf(request)
    file_path2 = generate_inward_report_daily_pdf(request)

    # Create the email
    email = EmailMessage(
        subject='Outward Report PDF',
        body='Please find the attached outward report in PDF format.',
        from_email='rradailyupdates@gmail.com',
        # to=['varad@ravirajanodisers.com', 'ravi@ravirajanodisers.com', 'pratikgosavi654@gmail.com', 'raj@ravirajanodisers.com'],
        to=['pratikgosavi654@gmail.com'],
    )

    # Attach the generated PDF
    email.attach_file(file_path)
    email.attach_file(file_path2)

    # Send the email
    email.send()

    return HttpResponse("Email with PDF attachment sent successfully.")




def generate_outward_report_pdf_email(request):
    # Create an in-memory buffer for the PDF file
    
   
    # Create an in-memory buffer for the PDF file
    buffer = BytesIO()

    # Create a PDF document with minimal margins
    pdf = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )
    
    elements = []

    # Define styles
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    title_style.alignment = 1  # Centered

    # Title of the PDF
    title = Paragraph("Material Outward Report", title_style)
    elements.append(title)

    # Query data from the database
    projects = project.objects.all().order_by("-id")

     # Check if request.GET is empty
    if is_invalid_get_params(request.GET):
        # Create a mutable copy of request.GET
        default_params = QueryDict(mutable=True)
        
        # Set default parameters

        default_params.update({
        'from_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0) - timedelta(days=1),
        'to_date_time': datetime.now().replace(hour=18, minute=0, second=0, microsecond=0)
        })

    else:

        default_params = request.GET
                
    project_filter_data = project_filter(default_params, queryset=project.objects.all())
    production_filter = project_matarial_production_filter(default_params, queryset=project_matarial_production.objects.all())


    print(default_params)

    # Apply production filter and prefetch the filtered production data
    project_qs = project_filter_data.qs  # Ensure this is a queryset
    production_qs = production_filter.qs  # Ensure this is a queryset

    print(production_qs)
    print(project_qs)

    # Prefetch the production data using the correct related_name for the relationship
    projects = project_qs.filter(
        project_production_n__in=production_qs  # Ensures only projects with related filtered production records
    ).prefetch_related(
        Prefetch('project_production_n', queryset=production_qs)  # Prefetch the filtered production data
    ).distinct()



      
        # Update the parameters
      


    light_orange = colors.Color(1, 0.647, 0, alpha=1)  # RGB for light orange

    for project_instance in projects:
        # Data for the main table
        main_table_data = [
            ["#", "Date", "Project Order No", "RRA Invoice No", "Customer Name"],
            [
                str(project_instance.pk),
                str(project_instance.DC_date),
                str(project_instance.order_id),
                str(project_instance.rra_invoice_no),
                str(project_instance.customer),
            ]
        ]

        # Create the main table
        main_table = Table(main_table_data)
        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgreen),  # Header row background
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),  # Data row background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Borders for table cells
        ]))

        elements.append(main_table)

        # Add an empty line for separation
        elements.append(Paragraph(''))

        # Nested table data
        nested_table_data = [
            ["#", "Item Code", "Quantity"]
        ]

        # Add data rows for each production
        productions = project_instance.project_production_n.all()
        for counter, production in enumerate(productions, start=1):
            nested_table_data.append([
                str(counter),
                str(production.item_code),
                str(production.production_quantity),
            ])

        # Add totals for this project
        totals = project_instance.project_production_n.aggregate(
            total_quantity=Sum('production_quantity'),
        )
        total_quantity = totals['total_quantity'] or 0
        nested_table_data.append([
            "",
            "Project Total",
            str(total_quantity),
        ])

        # Create the nested table
        nested_table = Table(nested_table_data, colWidths=[40, 200, 100, 100])  # Adjusted width for Item Code
        nested_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), light_orange),  # Header row background
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),  # Header text color
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (1, 1), (-1, -1), colors.whitesmoke),  # Data row background
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Borders for table cells
        ]))

        # Center the nested table within the page
        elements.append(Paragraph(''))  # Add a blank line for separation
        elements.append(nested_table)

        # Add an extra empty line for spacing between projects
        elements.append(Paragraph(''))


        # Add a blank line
        elements.append(Spacer(1, 0.5 * cm))



    # Build the PDF into the buffer
    pdf.build(elements)

    file_path = os.path.join(settings.MEDIA_ROOT, 'outward_report.pdf')


    with open(file_path, 'wb') as f:
        f.write(buffer.getvalue())

    # Close the buffer
    buffer.close()

    return file_path







def send_outward_report_pdf_email(request):
    # Generate and save the PDF
    
    
    file_path = generate_outward_report_pdf_email(request)

    # Create the email
    email = EmailMessage(
        subject='Outward Report PDF',
        body='Please find the attached outward report in PDF format.',
        from_email='rradailyupdates@gmail.com',
        # to=['varad@ravirajanodisers.com', 'ravi@ravirajanodisers.com', 'pratikgosavi654@gmail.com', 'raj@ravirajanodisers.com'],
        to=['pratikgosavi654@gmail.com'],
    )

    # Attach the generated PDF
    email.attach_file(file_path)

    # Send the email
    email.send()

    return HttpResponse("Email with PDF attachment sent successfully.")


from django.http import FileResponse



def download_outward_report_pdf(request):
    # Generate and save the PDF
    
    
    
    
    file_path = generate_outward_report_pdf_email(request)  # Assuming this function generates the PDF and returns the file path
    
    # Open the file in binary mode for reading
    pdf_file = open(file_path, 'rb')
    
    # Return the file as a downloadable response
    response = FileResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="outward_report.pdf"'
    
    return response






@login_required(login_url='login')
def assign_matarial_qr(request, project_id):

    print('hereeeeeeeeeeeeee')
    print(request.POST)
    print('------------')

    project_instance = project.objects.get(id = project_id)


    if request.method == 'POST':

        

        material = request.POST.getlist('material_id[]')
        scan = request.POST.getlist('scanned_value[]')



        print(material)
        print(scan)

        for i, z in zip(material, scan):

            print('-----------------')

            pro = project_matarial_qr.objects.get(id = i)
            print(pro)
            product_qr_instance = product_qr.objects.get(id = z)
            print(product_qr_instance)

            pro.product_qr = product_qr_instance
            pro.save()
            print('-----------------')


        url = reverse('list_project')

        print(url)

        some_data_to_dump = {
            'status': 'done',
            'url': url,
        }


        return JsonResponse((some_data_to_dump), safe = False) 




       


    else:

        forms = project_Form(instance = project_instance)

        data_form = product_Form()

        data = project_material.objects.filter(project = project_instance)


        context = {
            'form': forms,
            'data_form': data_form,
            'data': data,
            'project_id': project_id,
        }
        return render(request, 'transactions/assign_material_qr.html', context)


def delete_assign_material(request, assign_material_id, project_id):

    project_instance = project.objects.get(id = project_id)
    project_material_instance = project_material.objects.get(id=assign_material_id)
    project_material_qr_instance = project_matarial_qr.objects.get(project_material = project_material_instance, project_material__project = project_instance)
    project_material_qr_instance.delete()

    desginer_name = project_instance.employee_name

    project_material_instance.delete()

    # Assuming 'project_id' is the correct keyword argument for your 'update_project' view.
    return redirect('add_project_designer', project_id=project_material_instance.project.id)


def show_scanner_assign_matarial_qr(request):

    return render(request, 'transactions/update_qr_scanner.html')




def update_assign_matarial_qr(request, product_qr_id):

    product_qr_instance = product_qr.objects.get(id = product_qr_id)

    print(request.POST)

    if request.method == "POST":

        a = request.POST.get('size')
        b = request.POST.get('used_size')
        c = request.POST.get('left_size')
        e = request.POST.get('move_to_scratch')
        project_id = request.POST.get('order_id')
        cutter_id = request.POST.get('cutter')

        print('------------------------')

        print(cutter_id)

        try:
            project_instance = project.objects.get(id = project_id)
        except project.DoesNotExist as e:
            msg = "Project with project id:" + project_id
            return JsonResponse({'status' : 'error', 'msg' : msg})

        print('1111111111')
        
        try:
            
            project_material.objects.get(project = project_instance, sheet_no = product_qr_instance.id)

        except project_material.DoesNotExist:
            msg = "Sheet is not assign to project with order id " + project_id + " does not exsisit"
            return JsonResponse({'status' : 'error', 'msg' : msg})

        print('22222222222222')


        size_instance1, new_generated_size1 = size.objects.get_or_create(name = a)

        if size_instance1 == None:

            size_instance1 = new_generated_size1

        print('333333333333333')

        size_instance2, new_generated_size2 = size.objects.get_or_create(name = b)

        if size_instance2 == None:

            size_instance2 = new_generated_size2


        size_instance3, new_generated_size3= size.objects.get_or_create(name = c)
        
        if size_instance3 == None:

            size_instance3 = new_generated_size3


        print('4444444444')


        project_matarial_qr_instance = project_matarial_qr.objects.filter(product_qr__id = product_qr_id).first()

        product_instance = project_matarial_qr_instance.project_material.product

        cutter_intance = cutter.objects.get(id = cutter_id)

        # Now, retrieve the related project_qr instance
        c = material_history.objects.create(product_qr = product_qr_instance, previous_size = size_instance1, used_size = size_instance2, left_size = size_instance3, project = project_instance, cutter = cutter_intance)
        product_instance_new, product_created_new = product.objects.get_or_create(category = product_instance.category, thickness = product_instance.thickness, size = size_instance3, grade = product_instance.grade)
        



        print('555555555555555')





        project_material_qr_instance = project_matarial_qr.objects.get(product_qr = product_qr_instance, project_material = project_matarial_qr_instance.project_material)
        

        project_material_qr_instance.is_cutting_done = True
        project_material_qr_instance.save()

        if product_instance_new == None:
            product_instance_new = product_created_new


        print('666666666666666666666666')

        if e == "true":

            print(' in scratch ')


            if product_qr_instance.moved_to_left_over != True:


                print('----------------------------------2----------------------')

                print(product_instance.id)
               
                
                product_qr_instance.moved_to_scratch = True
                product_qr_instance.moved_to_left_over = False
                product_qr_instance.save()



                check_low_stock = product_qr.objects.filter(
                    moved_to_scratch=False,
                    moved_to_left_over=False,
                    product__isnull=False, product = product_qr_instance.product
                ).values(
                    'product',
                ).annotate(total_quantity=Count('id')).order_by('product').first()

                if check_low_stock:
                    total_quantity = check_low_stock['total_quantity']   # Handle None case
                else:
                    total_quantity = 0
             

                if total_quantity < 6:   #after moving this sheet -1 will happen so dont maake it 5
                    

                    print('-------------------1------------------')


             
                    
                    message_body =  "Category: " + str(product_qr_instance.product.category) + " Thickness: " +str(product_qr_instance.product.thickness)+ " Grade: " + str(product_qr_instance.product.grade) + " Quantity: " + str(total_quantity) + " " + " left."

                                    
                   
                    send_low_stock_notification(request, access_token, recipient_number, 'stock_alert', language_code, message_body)
                    

                    
              
            else:

                print('-------------------2------------------')

             

                product_qr_instance.moved_to_scratch = True
                product_qr_instance.save()




        else:

            print(' in left over ')


            
            if product_qr_instance.moved_to_left_over != True:
                
               
            
                check_low_stock = product_qr.objects.filter(
                    moved_to_scratch=False,
                    moved_to_left_over=False,
                    product__isnull=False, product = product_qr_instance.product
                ).values(
                    'product',
                ).annotate(total_quantity=Count('id')).order_by('product').first()

                if check_low_stock:
                    total_quantity = check_low_stock['total_quantity']   # Handle None case
                else:
                    total_quantity = 0
             

                if total_quantity < 6:   #after moving this sheet -1 will happen so dont maake it 5



                    print('------------------1----------------------')

                    

                    message_body =  "Category: " + str(product_qr_instance.product.category) + " Thickness: " +str(product_qr_instance.product.thickness)+ " Grade: " + str(product_qr_instance.product.grade) + " Quantity: " + str(total_quantity) + " " + " left."


                                    
                   
                    send_low_stock_notification(request, access_token, recipient_number, 'stock_alert', language_code, message_body)



                product_qr_instance.moved_to_left_over = True
                product_qr_instance.moved_to_scratch = False
                product_qr_instance.save()

           
             
        
        print('---------------')
        print(product_instance_new)
        print('---------------')

        product_qr_instance.product = product_instance_new
        product_qr_instance.save()
        

        # remove project for this qr
        return JsonResponse({'status' : 'done'})


    else:

        cutter_data = cutter.objects.all()

        context = {
            'data': product_qr_instance,
            'cutter_data': cutter_data,
            'product_qr_id' : product_qr_id
        }

        return render(request, 'transactions/update_assign_material_qr.html', context)







def get_cutting_values(request):

    project_id = request.GET.get('project_id')
    product_qr_id = request.GET.get('sheet_id')
    product_qr_instance = product_qr.objects.get(id = product_qr_id)

    data = project_material.objects.get(project = project_id, sheet_no = product_qr_instance.id)

    mm_length = data.length / 25.4
    mm_width =  data.width / 25.4
    used_sqinch = round(mm_length * mm_width, 2) 



    # Round to 2 decimal places
    used_sqinch = round(used_sqinch, 2)

    data = {
        'length': data.length,  # Replace with actual field names
        'width': data.width,    # Replace with actual field names
        'used_sqinch': used_sqinch  # Example calculation
    }


    

    return JsonResponse(data)


def delete_matarial_history(request, history_id):

    pass

    
from django.forms import modelformset_factory

def add_production(request, project_id):

    project_instance = project.objects.get(id = project_id)

    if request.method == "POST":

        quantity = request.POST.getlist('quantity[]')
        item_code_id = request.POST.getlist('item_code[]')
        material = request.POST.getlist('materialsId[]')
        production_id = request.POST.getlist('production_id[]')
        completed = request.POST.get('completed')


        for a, b, c, d in zip(production_id, item_code_id, quantity, material):

            material_instance = project_matarial_qr.objects.get(id = d)

            if a and a!= '0':

                project_material_instnace = project_matarial_production.objects.get(id = a)
                item_code_instance = item_code.objects.get(id = b)

                project_material_instnace.item_code = item_code_instance
                project_material_instnace.production_quantity = c
                
                project_material_instnace.save()

                print('here')

            else:
                    
                item_code_instance = item_code.objects.get(id = b)
                instance = project_matarial_production.objects.create(project_matarial_qr = material_instance, item_code = item_code_instance, production_quantity = c, project = project_instance)

        if completed == "true":
            project_instance.completed = True
            project_instance.save()

        return JsonResponse({'status' : 'done'})
        
    else:

        data = project_material.objects.filter(project = project_instance)

        item_code_data = item_code.objects.all()

        print(item_code_data)

        # Create the formset with the initial data
       
        context = {
            'form': project_Form(instance=project_instance),
            'data' : data,
            'item_code_data' : item_code_data,
            'project_material_form' : project_matarial_qr_Form(),
            'project_id' : project_id,
        }

        return render(request,'transactions/add_production.html', context )



def delete_production_entry(request, project_id, production_id):


    project_matarial_production.objects.get(id = production_id).delete()

    url = reverse('update_project_accountant', args=[project_id])
    return redirect(url)






import qrcode

from django.core.files.base import ContentFile
from io import BytesIO


import qrcode
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, Image, Spacer
from io import BytesIO


def generate_qr_codes_pdf(qr_codes):
     
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    col_count = 4
    qr_width = 150
    qr_height = 150

    data = []
    row = []
    for qr in qr_codes:
        qr_buffer = BytesIO()
        qr.save(qr_buffer, format='PNG')
        qr_image = Image(qr_buffer)
        qr_image.drawWidth = qr_width
        qr_image.drawHeight = qr_height
        row.append(qr_image)

        if len(row) == col_count:
            data.append(row)
            row = []

    # Add any remaining QR codes to the last row
    if row:
        data.append(row)

    table = Table(data)
    elements.append(table)

    doc.build(elements)

    return buffer


def generate_product_qr(request):

    quantity = request.POST.get('quantity')
    qr_codes = []

    # Specify the desired size in pixels (e.g., 600x600)
    qr_size = 600

    for i in range(int(quantity)):
        
        a = product_qr.objects.create()
        product_qr_shelf.objects.create(product_qr = a)

        # Adjust the box_size based on the desired size
        box_size = qr_size // 4  # You can experiment with different values

        qr = qrcode.QRCode(
            version=1,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=box_size,
            border=0,
        )
        qr.add_data(a.id)
        qr.make(fit=True)

        img = qr.make_image(fill_color="black", back_color="white", size=(qr_size, qr_size))
        buffer = BytesIO()
        img.save(buffer, format="PNG")

        qr_code_image = ContentFile(buffer.getvalue())

        a.qr_code.save(f"qr_code_{a.id}.png", qr_code_image)

        qr_codes.append(a)

    context ={
        'data' : qr_codes
    }

    return render(request, 'transactions/html_qr.html', context)
   

def generate_product_qr_with_values(request):

    if request.method == "POST":

        quantity = request.POST.get('quantity')
        shelf_id = request.POST.get('shelf')
        supplier_id = request.POST.get('supplier')
        category_id = request.POST.get('category')
        size_id = request.POST.get('size')
        thickness_id = request.POST.get('thickness')
        grade_id = request.POST.get('grade')
        finish_id = request.POST.get('finish')
        date_of_pur_id = request.POST.get('date_of_pur')


        category_instance = category.objects.get(id = category_id)
        size_instance = size.objects.get(id = size_id)
        thickness_instance = thickness.objects.get(id = thickness_id)
        grade_instance = grade.objects.get(id = grade_id)

        book, created = product.objects.get_or_create(category = category_instance, size =  size_instance, thickness = thickness_instance, grade = grade_instance)
        
        if book:

            product_instance = book

        else:

            product_instance = created

        qr_codes = []

        # Specify the desired size in pixels (e.g., 600x600)
        qr_size = 600

        for i in range(int(quantity)):
            
            

            supplier_instance= dealer.objects.get(id = supplier_id)

            a = product_qr.objects.create(product = product_instance, date_of_pur = date_of_pur_id, finish = finish_id, supplier = supplier_instance)
            
            shelf_instance = shelf.objects.get(id = shelf_id)
            
            product_qr_shelf.objects.create(product_qr = a, shelf = shelf_instance)

            # Adjust the box_size based on the desired size
            box_size = qr_size // 4  # You can experiment with different values

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=box_size,
                border=0,
            )
            qr.add_data(a.id)
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white", size=(qr_size, qr_size))
            buffer = BytesIO()
            img.save(buffer, format="PNG")

            qr_code_image = ContentFile(buffer.getvalue())

            a.qr_code.save(f"qr_code_{a.id}.png", qr_code_image)

            qr_codes.append(a)

        context ={
            'data' : qr_codes
        }

        return render(request, 'transactions/html_qr.html', context)
    else:

        print('-----------------')
        print('-----------------')
        print('-----------------')
        print('-----------------')
        print('-----------------')

        form = product_Form()
        form_qr = product_qr_Form()
        product_qr_shelf_form = product_qr_shelf_Form()


        context ={
            'form' : form,
            'form_qr': form_qr,
            'product_qr_shelf_form': product_qr_shelf_form,

        }

        return render(request, 'transactions/generate_product_qr_with_values.html', context)


def from_to_generate_product_qr(request):


    from_no = request.POST.get('from')
    to_no = request.POST.get('to')

    qr_codes = []

    # Specify the desired size in pixels (e.g., 600x600)
    qr_size = 600

    for i in range(int(from_no), int(to_no) + 1):
        
        try:
            a = product_qr.objects.get(id = i)
        except product_qr.DoesNotExist:
            a = None
        
        if a:
            # Adjust the box_size based on the desired size
            box_size = qr_size // 4  # You can experiment with different values

            qr = qrcode.QRCode(
                version=1,
                error_correction=qrcode.constants.ERROR_CORRECT_L,
                box_size=box_size,
                border=0,
            )
            qr.add_data(a.id)
            qr.make(fit=True)

            img = qr.make_image(fill_color="black", back_color="white", size=(qr_size, qr_size))
            buffer = BytesIO()
            img.save(buffer, format="PNG")

            qr_code_image = ContentFile(buffer.getvalue())

            a.qr_code.save(f"qr_code_{a.id}.png", qr_code_image)

            qr_codes.append(a)

    context ={
        'data' : qr_codes
    }

    return render(request, 'transactions/html_qr.html', context)
  




from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Image


def print_single_qr(request, product_qr_id):

    a = product_qr.objects.get(id = product_qr_id)


    qr_size = 600

    box_size = qr_size // 4  # You can experiment with different values

    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=box_size,
        border=0,
    )
    qr.add_data(a.id)
    qr.make(fit=True)

    img = qr.make_image(fill_color="black", back_color="white", size=(qr_size, qr_size))
    buffer = BytesIO()
    img.save(buffer, format="PNG")

    qr_code_image = ContentFile(buffer.getvalue())

    a.qr_code.save(f"qr_code_{a.id}.png", qr_code_image)

    

    return render(request, 'transactions/html_qr_single.html', {'data' : a})

def print_label(request, project_id, product_qr_id):

    project_instance = project.objects.get(id = project_id)

    data = project_material.objects.get(project = project_id, sheet_no = product_qr_id)

   
    context = {
        
        'project_id': project_id,
        'product_qr_id': product_qr_id,
        # 'sqinch_alloted': product_qr_id,
        'date': datetime.now(),
        'project_name': project_instance.customer,
        'cutting_size': data,
    }

    

    return render(request, 'transactions/html_print_label.html', {'data' : context})

def list_generated_product_qr(request):

    data = product_qr.objects.all().order_by("-id")

    status = request.GET.get('status')
    sheet_id = request.GET.get('sheet_id')

    if sheet_id:
        data = data.filter(id = sheet_id)

    
    elif status:
        data = data.filter(status = status)

    
    page = request.GET.get('page', 1)
    paginator = Paginator(data, 50)

    try:
        data = paginator.page(page)
    except PageNotAnInteger:
        data = paginator.page(1)
    except EmptyPage:
        data = paginator.page(paginator.num_pages)

    context = {
        
        'data': data,
        'status': status,
        'sheet_id': sheet_id,
    }

    return render(request, 'transactions/list_all_generated_qr.html', context)



@login_required(login_url='login')
def sheet_report(request):

    filter_set = product_qr_filter(request.GET, queryset=product_qr.objects.all())
    data = filter_set.qs

    sheet_count = data.count()

   
    context = {
        
        'sheet_filter': filter_set,
        'data': data,
        'sheet_count': sheet_count,
    }

    return render(request, 'transactions/sheet_report.html', context)




def sheet_report_downlaod(request):

    filter_set = product_qr_filter(request.GET, queryset=product_qr.objects.all())
    data = filter_set.qs

    
    order_filters_data1 = list(data.values_list('id', 'product__category', 'product__size', 'product__grade', 'product__thickness', 'finish', 'date', 'moved_to_left_over', 'moved_to_scratch', 'product__size'))
    order_filters_data = list(map(list, order_filters_data1))
    

    vals = []
        
    vals1 = []

    
    vals.append([''])
    vals.append(['SHEETS REPORT'])
    vals.append([''])
    vals.append([''])
    vals1.append("Sr No")
    vals1.append("Sheet Id")
    vals1.append("Matrial Name")
    vals1.append("Size")
    vals1.append("Grade")
    vals1.append("Thickness")
    vals1.append("Finish")
    vals1.append("Date")
    vals1.append("New Sheet")
    vals1.append("Leftover Sheet")
    vals1.append("Dead Sheet")
    vals1.append("Left SQinch")
    vals.append(vals1)

    counteer = 1

    for i in order_filters_data:
        vals1 = []
        vals1.append(counteer)
        counteer += 1
        vals1.append(i[0])  # Sheet Id
        vals1.append(i[1])  # Sheet Id
        vals1.append(i[2])  # Sheet Id
        vals1.append(i[3])  # Sheet Id
        vals1.append(i[4])  # Sheet Id
        vals1.append(i[5])  # Sheet Id
        vals1.append(i[6])  # Sheet Id
        
        # Conditional logic for "New Sheet", "Leftover Sheet", and "Dead Sheet"
        if i[7]:  # moved_to_left_over is True
            vals1.append("No")  # New Sheet
            vals1.append("Yes")  # Leftover Sheet
            vals1.append("No")  # Dead Sheet
        elif i[8]:  # moved_to_scratch is True
            vals1.append("No")  # New Sheet
            vals1.append("No")  # Leftover Sheet
            vals1.append("Yes")  # Dead Sheet
        else:  # Neither moved_to_left_over nor moved_to_scratch is True
            vals1.append("Yes")  # New Sheet
            vals1.append("No")  # Leftover Sheet
            vals1.append("No")  # Dead Sheet
        
        vals.append(vals1)


    name = "Sales_Report.csv"
    path = os.path.join(BASE_DIR) + '\static\csv\\' + name

    
    with open(path,  'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(vals)


        link = os.path.join(BASE_DIR) + '\static\csv\\' + name

    with open(path,  'r', newline="") as f:
        mime_type  = mimetypes.guess_type(link)

        response = HttpResponse(f.read(), content_type=mime_type)
        response['Content-Disposition'] = 'attachment;filename=' + str(link)

        return response



def sheet_status_report(request):

    data = product_qr.objects.all().order_by("id")



    
    order_filters_data1 = list(data.values_list('id', 'moved_to_left_over', 'moved_to_scratch', 'product__size'))
    order_filters_data = list(map(list, order_filters_data1))
    

    vals = []
        
    vals1 = []

    
    vals.append([''])
    vals.append(['SHEETS REPORT'])
    vals.append([''])
    vals.append([''])
    
    vals1.append("Sr No")
    vals1.append("Sheet Id")
    vals1.append("New Sheet")
    vals1.append("Leftover Sheet")
    vals1.append("Dead Sheet")
    vals1.append("Left SQinch")
    vals.append(vals1)

    counteer = 1

    for i in order_filters_data:
        vals1 = []
        vals1.append(counteer)
        counteer += 1
        vals1.append(i[0])  # Sheet Id
        
        # Conditional logic for "New Sheet", "Leftover Sheet", and "Dead Sheet"
        if i[1]:  # moved_to_left_over is True
            vals1.append("No")  # New Sheet
            vals1.append("Yes")  # Leftover Sheet
            vals1.append("No")  # Dead Sheet
        elif i[2]:  # moved_to_scratch is True
            vals1.append("No")  # New Sheet
            vals1.append("No")  # Leftover Sheet
            vals1.append("Yes")  # Dead Sheet
        else:  # Neither moved_to_left_over nor moved_to_scratch is True
            vals1.append("Yes")  # New Sheet
            vals1.append("No")  # Leftover Sheet
            vals1.append("No")  # Dead Sheet
        
        vals1.append(i[3])  # Left SQinch
        vals.append(vals1)


    name = "Sales_Report.csv"
    path = os.path.join(BASE_DIR) + '\static\csv\\' + name

    
    with open(path,  'w', newline="") as f:
        writer = csv.writer(f)
        writer.writerows(vals)


        link = os.path.join(BASE_DIR) + '\static\csv\\' + name

    with open(path,  'r', newline="") as f:
        mime_type  = mimetypes.guess_type(link)

        response = HttpResponse(f.read(), content_type=mime_type)
        response['Content-Disposition'] = 'attachment;filename=' + str(link)

        return response



def scanner_page(request):
    
    product_id = request.POST.get('scanned_value')

    context = {
                'product_id': product_id,
        }

    return render(request, 'transactions/scanner.html', context)

def assign_values_to_qr_page(request):

    scanned_value = request.POST.get("scanned_value")

    
    redirect_url = reverse('assign_values_to_qr', args=[scanned_value]) 
    response_data = {'status' : 'success', 'redirect_url': redirect_url}
    return JsonResponse(response_data)
    

def assign_values_to_qr_page_data(request, product_qr_id):

    product_qr_instance = product_qr.objects.get(id = product_qr_id)

    product_form = product_Form(instance = product_qr_instance.product)
    
    context = {
        'form': product_form,
    }

    return render(request, 'transactions/assign_values_to_qr_page_data.html', context)
    

from django.core.files.storage import FileSystemStorage

import copy



def assign_values_to_qr(request, product_qr_id):
    
    print('hereeeeeeee')
    print(product_qr_id)
    product_id = product_qr_id

    try:
        product_qr_instance = product_qr.objects.get(id=product_id)
    except product_qr.DoesNotExist as e:
        # Handle the error and construct an error response
        error_message = str(e)
        return JsonResponse({'status': 'error', 'message': error_message}, status=400)
    
    if request.method == 'POST':

        product_qr_old = product_qr_old = copy.copy(product_qr_instance)

        print('here checking')
        print(product_qr_old.is_fix)

       
        if product_qr_instance.moved_to_left_over == False and product_qr_instance.moved_to_scratch == False:
            print('1111')


            category_id = request.POST.get('category')
            size_id = request.POST.get('size')
            thickness_id = request.POST.get('thickness')
            grade_id = request.POST.get('grade')

            print('printing data ------------------')
           
            print('printing data ------------------')

            category_instance = category.objects.get(id = category_id)
            size_instance = size.objects.get(id = size_id)
            thickness_instance = thickness.objects.get(id = thickness_id)
            grade_instance = grade.objects.get(id = grade_id)

            book, created = product.objects.get_or_create(category = category_instance, size =  size_instance, thickness = thickness_instance, grade = grade_instance)

            print('2222')

            form2 = product_qr_Form(request.POST, request.FILES, instance = product_qr_instance)
            if form2.is_valid():


                print('valid')


                # Save the file to a specific location using FileSystemStorage or another storage backend.
                form2.save()
                shelf_id = request.POST.get('shelf')
                print('----------')
                print(shelf_id)
                print('----------')


                shelf_instance = shelf.objects.get(id = shelf_id)
                product_qr_shelf_instance = product_qr_shelf.objects.get(product_qr = product_qr_instance)
                product_qr_shelf_instance.shelf = shelf_instance
                product_qr_shelf_instance.save()
            
                print('valid2')

                if book:

                    product_instance = book

                else:

                    product_instance = created

                messages.success(request, 'Values added successfully')
                print('valid3')

                
                product_qr_instance.product = product_instance
                product_qr_instance.is_fix = True
                product_qr_instance.save()

                redirect_url = reverse('list_generated_product_qr')

                return redirect(redirect_url)

                
            else:

                print(form.errors) 
        else:

            messages.error(request, 'Material already been used please contact pratik')

            return redirect('assign_values_to_qr', product_qr_id=product_qr_id)

    else:    

        form = product_Form(instance=product_qr_instance.product)
        form_qr = product_qr_Form(instance=product_qr_instance)

        print(product_qr_id)
        data = material_history.objects.filter(product_qr__id = product_qr_id)

        product_qr_shelf_instance = product_qr_shelf.objects.get(product_qr = product_qr_instance)

        product_qr_shelf_form = product_qr_shelf_Form(instance = product_qr_shelf_instance)


        context = {
            'form': form,
            'form_qr': form_qr,
            'data': data,
            'sheet_no': product_qr_id,
            'product_qr_id': product_qr_id,
            'project_filter': project_filter(),

            'product_qr_shelf_form': product_qr_shelf_form,
        }

        return render(request, 'transactions/assign_value_to_qr.html', context)
    
def delete_history(request, history_id):

    material_history_instance = material_history.objects.get(id = history_id)

    product_qr_instance = material_history_instance.product_qr

    product_old = copy.copy(product_qr_instance.product)

    obj, created = product.objects.get_or_create(category_id = product_qr_instance.product.category.id, size_id = material_history_instance.previous_size, grade_id = product_qr_instance.product.grade.id, thickness_id = product_qr_instance.product.thickness.id)

    if obj:
        product_qr_instance.product = obj
        product_qr_instance.save()
    else:
        product_qr_instance.product = created
        product_qr_instance.save()


    if product_qr_instance.moved_to_scratch:

        print('runung 1')

        material_history_count = material_history.objects.filter(product_qr = product_qr_instance).count()
        
        instance_previous_stock = scratch_stock.objects.get(product = product_old)
        instance_previous_stock.quantity = instance_previous_stock.quantity - 1
        instance_previous_stock.save()
        
        if material_history_count == 1:

            print('22')  

            instance, created = stock.objects.get_or_create(product = product_qr_instance.product)

            if instance:

                instance.quantity = instance.quantity + 1
                instance.save()

            else:

                created.quantity = 1
                created.save()

            product_qr_instance.moved_to_scratch = False
            product_qr_instance.moved_to_left_over = False
            product_qr_instance.save()

        else:

            print('11')  
            

            product_qr_instance.moved_to_scratch = False
            product_qr_instance.moved_to_left_over = True
            product_qr_instance.save()

            instance, created = left_over_stock.objects.get_or_create(product = product_qr_instance.product)

            if instance:

                instance.quantity = instance.quantity + 1
                instance.save()

            else:

                created.quantity = 1
                created.save()

    


    elif product_qr_instance.moved_to_left_over:

        print('runung 2')

        material_history_count = material_history.objects.filter(product_qr = product_qr_instance).count()

        instance_previous_stock = left_over_stock.objects.get(product = product_old)
        instance_previous_stock.quantity = instance_previous_stock.quantity - 1
        instance_previous_stock.save()

        if material_history_count == 1:

            instance, created = stock.objects.get_or_create(product = product_qr_instance.product)

            if instance:

                instance.quantity = instance.quantity + 1
                instance.save()

            else:

                created.quantity = 1
                created.save()

            product_qr_instance.moved_to_scratch = False
            product_qr_instance.moved_to_left_over = False
            product_qr_instance.save()

        else:
            
            instance, created = left_over_stock.objects.get_or_create(product = product_qr_instance.product)

            if instance:

                instance.quantity = instance.quantity + 1
                instance.save()

            else:

                created.quantity = 1
                created.save()



    material_history_instance.delete()

    return redirect('list_generated_product_qr')



def check_sheet(request):

    data = product_qr.objects.all()

    for i in data:

        if i.moved_to_left_over == "None"  and i.moved_to_scratch == "None":
            if i.product:
                print(i.id, ' ', i.moved_to_left_over, ' ', i.moved_to_scratch, ' ', 'True', ' ', i.product.size)
        else:
            if i.product:
                print(i.id, ' ', i.moved_to_left_over, ' ', i.moved_to_scratch, ' ', 'False', ' ', i.product.size)



def assign_values_to_qr_page(request):

    scanned_value = request.POST.get("scanned_value")


    redirect_url = reverse('assign_values_to_qr', args=[scanned_value]) 
    response_data = {'status' : 'success', 'redirect_url': redirect_url}
    return JsonResponse(response_data)


def update_product(request, product_id):

    pass

def verify_password(request):

    
    emplpyee_id = request.POST.get("employee")
    password = request.POST.get("password")

    employee_instance = employee.objects.get(id = emplpyee_id)

    if employee_instance.password == password:

        return JsonResponse({'status' : "done"})
    
    else:

        return JsonResponse({'status' : "wrong"})



import openpyxl

def demo(request):

    logo_path = os.path.join(settings.BASE_DIR, 'static', 'csv.xlsx')  # Replace with the actual path to your logo



    workbook = openpyxl.load_workbook(logo_path)
    sheet = workbook.active  # Use the first sheet or specify the sheet name

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        item_code, description = row[0], row[1]  # Adjust indexes if necessary
        
        # Create or update the record
        inward_item_code.objects.update_or_create(
            item_code=item_code,
            defaults={'description': description}
        )
